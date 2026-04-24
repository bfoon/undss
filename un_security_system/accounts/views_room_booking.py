from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from django.urls import reverse_lazy, reverse
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from django.http import JsonResponse, HttpResponse, Http404
import qrcode
import io
from django.views.decorators.http import require_POST, require_GET
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, timedelta, date as date_cls
from datetime import datetime, date, time
import threading
from django.shortcuts import render

from django.conf import settings
from django.core.mail import send_mail, EmailMessage, EmailMultiAlternatives
from .utils import generate_booking_ics
from django.core.paginator import Paginator
from django.db.models import Q, Count, Prefetch
from datetime import date
from django.db import transaction, IntegrityError
from datetime import datetime, time as time_cls, timedelta, date as date_cls
from django.utils.text import slugify
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image, ImageDraw, ImageFont

from .models import (
    Room,
    RoomBooking,
    RoomApprover,
    RoomBookingSeries,
    MeetingAttendee,
    AttendanceRecord,
)
from .forms import (
    RoomBookingForm,
    RoomBookingApprovalForm,
    RoomForm,
    RoomSeriesApprovalForm,
    MeetingAttendeeForm,
)
# ICT focal lookup — graceful so the file works if AgencyAssetRoles isn't present.
try:
    from .models import AgencyAssetRoles, User as _User

    _HAS_AGENCY_ROLES = True
except ImportError:
    _HAS_AGENCY_ROLES = False


# ======================= EMAIL HELPERS =======================

def _send_email_async(subject, message, recipients):
    """
    Send email in the background using a thread.
    """
    recipients = [e for e in recipients if e]  # filter empty
    if not recipients:
        return

    from_email = getattr(settings, "DEFAULT_FROM_EMAIL", None)
    if not from_email:
        return

    def _task():
        try:
            send_mail(subject, message, from_email, recipients, fail_silently=True)
        except Exception:
            # Optionally log here
            pass

    # fire-and-forget background thread
    threading.Thread(target=_task, daemon=True).start()

def ensure_booking_registration_code(booking):
    """
    Ensure a booking has a registration code when invite link is enabled.
    Returns the booking.registration_code after saving if needed.
    """
    if booking.enable_invite_link and not booking.registration_code:
        import uuid
        booking.registration_code = uuid.uuid4()
        booking.save(update_fields=["registration_code"])
    return booking.registration_code

def notify_approvers_new_booking(booking):
    approver_emails = list(
        RoomApprover.objects.filter(
            room=booking.room,
            is_active=True,
        ).select_related("user").values_list("user__email", flat=True)
    )

    requester_name = booking.requested_by.get_full_name() or booking.requested_by.username
    subject = f"[Room Booking] New request for {booking.room.name} on {booking.date}"
    message = (
        f"Dear Approver,\n\n"
        f"A new room booking request has been submitted for:\n\n"
        f"  Room: {booking.room.name} ({booking.room.code})\n"
        f"  Date: {booking.date}\n"
        f"  Time: {booking.start_time} – {booking.end_time}\n"
        f"  Title: {booking.title}\n"
        f"  Requested by: {requester_name}\n\n"
        f"Please log into the booking system and review this request under 'My Approvals'.\n\n"
        f"Thank you."
    )

    _send_email_async(subject, message, approver_emails)


def notify_approvers_new_series(series):
    """Notify approvers about a new recurring booking series"""
    approver_emails = list(
        RoomApprover.objects.filter(
            room=series.room,
            is_active=True,
        ).select_related("user").values_list("user__email", flat=True)
    )

    requester_name = series.requested_by.get_full_name() or series.requested_by.username
    freq_display = series.get_frequency_display() if series.frequency else "One-time"
    occurrence_count = series.occurrences.count()

    subject = f"[Room Booking] New recurring series for {series.room.name}"
    message = (
        f"Dear Approver,\n\n"
        f"A new RECURRING room booking series has been submitted:\n\n"
        f"  Room: {series.room.name} ({series.room.code})\n"
        f"  Title: {series.title}\n"
        f"  Frequency: {freq_display}\n"
        f"  Start Date: {series.start_date}\n"
        f"  End Date: {series.end_date or 'No end date'}\n"
        f"  Time: {series.start_time} – {series.end_time}\n"
        f"  Number of occurrences: {occurrence_count}\n"
        f"  Requested by: {requester_name}\n\n"
        f"Please log into the booking system and review this series under 'My Approvals'.\n"
        f"You can approve/reject the entire series at once.\n\n"
        f"Thank you."
    )

    _send_email_async(subject, message, approver_emails)


def notify_requester_series_submitted(series):
    if not series.requested_by.email:
        return

    occurrence_count = series.occurrences.count()
    freq_display = series.get_frequency_display() if series.frequency else "One-time"

    subject = f"[Room Booking] Recurring series submitted for {series.room.name}"
    message = (
        f"Dear {series.requested_by.get_full_name() or series.requested_by.username},\n\n"
        f"Your recurring booking series has been submitted and is pending approval.\n\n"
        f"Details:\n"
        f"  Room: {series.room.name} ({series.room.code})\n"
        f"  Title: {series.title}\n"
        f"  Frequency: {freq_display}\n"
        f"  Start Date: {series.start_date}\n"
        f"  End Date: {series.end_date or 'No end date'}\n"
        f"  Time: {series.start_time} – {series.end_time}\n"
        f"  Number of occurrences: {occurrence_count}\n"
        f"  Status: Pending approval\n\n"
        f"You will receive another email once your series is approved or rejected.\n\n"
        f"Thank you."
    )

    _send_email_async(subject, message, [series.requested_by.email])


def notify_requester_series_approved(series):
    if not series.requested_by.email:
        return

    approver_name = (
        series.approved_by.get_full_name() or series.approved_by.username
        if series.approved_by
        else "System (auto-approved)"
    )
    occurrence_count = series.occurrences.count()

    subject = f"[Room Booking] Recurring series APPROVED: {series.room.name}"
    message = (
        f"Dear {series.requested_by.get_full_name() or series.requested_by.username},\n\n"
        f"Your recurring room booking series has been APPROVED.\n\n"
        f"Details:\n"
        f"  Room: {series.room.name} ({series.room.code})\n"
        f"  Title: {series.title}\n"
        f"  Start Date: {series.start_date}\n"
        f"  End Date: {series.end_date or 'No end date'}\n"
        f"  Time: {series.start_time} – {series.end_time}\n"
        f"  Number of occurrences: {occurrence_count}\n"
        f"  Approved by: {approver_name}\n\n"
        f"All {occurrence_count} bookings in this series have been approved.\n\n"
        f"Please remember to leave the room neat and tidy as you found it.\n\n"
        f"Thank you."
    )

    _send_email_async(subject, message, [series.requested_by.email])


def notify_requester_series_rejected(series):
    if not series.requested_by.email:
        return

    approver_name = series.approved_by.get_full_name() if series.approved_by else "Approver"

    subject = f"[Room Booking] Recurring series REJECTED: {series.room.name}"
    message = (
        f"Dear {series.requested_by.get_full_name() or series.requested_by.username},\n\n"
        f"Your recurring room booking series has been REJECTED.\n\n"
        f"Details:\n"
        f"  Room: {series.room.name} ({series.room.code})\n"
        f"  Title: {series.title}\n"
        f"  Start Date: {series.start_date}\n"
        f"  End Date: {series.end_date or 'No end date'}\n"
        f"  Time: {series.start_time} – {series.end_time}\n"
        f"  Rejected by: {approver_name}\n\n"
        f"Reason provided:\n"
        f"{series.rejection_reason or 'No reason provided.'}\n\n"
        f"If you have questions, please contact the approver or the ICT/administration team.\n\n"
        f"Thank you."
    )

    _send_email_async(subject, message, [series.requested_by.email])


def notify_requester_booking_submitted(booking):
    if not booking.requested_by.email:
        return

    subject = f"[Room Booking] Request submitted for {booking.room.name} on {booking.date}"
    message = (
        f"Dear {booking.requested_by.get_full_name() or booking.requested_by.username},\n\n"
        f"Your booking request has been submitted and is pending approval.\n\n"
        f"Details:\n"
        f"  Room: {booking.room.name} ({booking.room.code})\n"
        f"  Date: {booking.date}\n"
        f"  Time: {booking.start_time} – {booking.end_time}\n"
        f"  Title: {booking.title}\n"
        f"  Status: Pending approval\n\n"
        f"You will receive another email once your request is approved or rejected.\n\n"
        f"Thank you."
    )

    _send_email_async(subject, message, [booking.requested_by.email])



def notify_requester_booking_approved(request, booking):
    if not booking.requested_by.email:
        return

    registration_link = None
    attendance_link = None

    if booking.enable_invite_link:
        ensure_booking_registration_code(booking)
        if booking.enable_attendance:
            attendance_link = request.build_absolute_uri(
                reverse('accounts:meeting_attendance_page', args=[booking.registration_code])
            )
        else:
            registration_link = request.build_absolute_uri(
                reverse('accounts:meeting_registration', args=[booking.registration_code])
            )

    requester_name = booking.requested_by.get_full_name() or booking.requested_by.username

    message = (
        f"Dear {requester_name},\n\n"
        f"Your booking for '{booking.title}' has been APPROVED.\n\n"
        f"Details:\n"
        f"  Room: {booking.room.name} ({booking.room.code})\n"
        f"  Date: {booking.date}\n"
        f"  Time: {booking.start_time} \u2013 {booking.end_time}\n\n"
    )

    if attendance_link:
        message += (
            f"Attendance QR / Check-in Link:\n{attendance_link}\n\n"
            "Display or share this QR code so attendees can scan it on the day to check in.\n\n"
        )
    elif registration_link:
        message += (
            f"Attendee Registration Link:\n{registration_link}\n\n"
            "Please share this link with your attendees so they can register in advance.\n\n"
        )

    message += (
        "A calendar invite has been attached to this email.\n\n"
        "Please remember to leave the room neat and tidy as you found it.\n\n"
        "Thank you."
    )

    _send_email_async(
        f"Approved: {booking.title}",
        message,
        [booking.requested_by.email]
    )

def notify_attendee_of_registration(attendee):
    subject = f"Registration Confirmed: {attendee.booking.title}"
    message = (
        f"Dear {attendee.name},\n\n"
        f"Thank you for registering for the meeting: '{attendee.booking.title}'.\n\n"
        f"Date: {attendee.booking.date.strftime('%A, %B %d, %Y')}\n"
        f"Time: {attendee.booking.start_time.strftime('%I:%M %p')}\n"
        f"Room: {attendee.booking.room.name}\n\n"
        "We look forward to seeing you."
    )
    _send_email_async(subject, message, [attendee.email])

def notify_requester_booking_rejected(booking):
    if not booking.requested_by.email:
        return

    approver_name = booking.approved_by.get_full_name() if booking.approved_by else "Approver"
    subject = f"[Room Booking] Rejected: {booking.room.name} on {booking.date}"
    message = (
        f"Dear {booking.requested_by.get_full_name() or booking.requested_by.username},\n\n"
        f"Your room booking request has been REJECTED.\n\n"
        f"Details:\n"
        f"  Room: {booking.room.name} ({booking.room.code})\n"
        f"  Date: {booking.date}\n"
        f"  Time: {booking.start_time} – {booking.end_time}\n"
        f"  Title: {booking.title}\n"
        f"  Rejected by: {approver_name}\n\n"
        f"Reason provided:\n"
        f"{booking.rejection_reason or 'No reason provided.'}\n\n"
        f"If you have questions, please contact the approver or the ICT/administration team.\n\n"
        f"Thank you."
    )

    _send_email_async(subject, message, [booking.requested_by.email])


def notify_approvers_booking_cancelled(booking):
    """Notify approvers when an approved booking is cancelled."""
    approver_emails = list(
        RoomApprover.objects.filter(
            room=booking.room,
            is_active=True,
        ).select_related("user").values_list("user__email", flat=True)
    )

    requester_name = booking.requested_by.get_full_name() or booking.requested_by.username
    subject = f"[Room Booking] Cancelled: {booking.room.name} on {booking.date}"
    message = (
        f"Dear Approver,\n\n"
        f"A previously approved room booking has been CANCELLED by the requester.\n\n"
        f"Details:\n"
        f"  Room: {booking.room.name} ({booking.room.code})\n"
        f"  Date: {booking.date}\n"
        f"  Time: {booking.start_time} – {booking.end_time}\n"
        f"  Title: {booking.title}\n"
        f"  Cancelled by: {requester_name}\n\n"
        f"The room is now available for this time slot.\n\n"
        f"Thank you."
    )

    _send_email_async(subject, message, approver_emails)


def notify_approvers_series_cancelled(series, occurrence_count):
    """Notify approvers when an approved recurring series is cancelled."""
    approver_emails = list(
        RoomApprover.objects.filter(
            room=series.room,
            is_active=True,
        ).select_related("user").values_list("user__email", flat=True)
    )

    requester_name = series.requested_by.get_full_name() or series.requested_by.username
    freq_display = series.get_frequency_display() if series.frequency else "One-time"

    subject = f"[Room Booking] Recurring series cancelled: {series.room.name}"
    message = (
        f"Dear Approver,\n\n"
        f"A previously approved recurring booking series has been CANCELLED by the requester.\n\n"
        f"Details:\n"
        f"  Room: {series.room.name} ({series.room.code})\n"
        f"  Title: {series.title}\n"
        f"  Frequency: {freq_display}\n"
        f"  Start Date: {series.start_date}\n"
        f"  End Date: {series.end_date or 'No end date'}\n"
        f"  Time: {series.start_time} – {series.end_time}\n"
        f"  Number of occurrences: {occurrence_count}\n"
        f"  Cancelled by: {requester_name}\n\n"
        f"All {occurrence_count} bookings in this series have been cancelled.\n"
        f"The room is now available for these time slots.\n\n"
        f"Thank you."
    )

    _send_email_async(subject, message, approver_emails)


# Email notification function for individual occurrence

def notify_approvers_occurrence_cancelled(occurrence):
    """Notify approvers when an individual series occurrence is cancelled."""
    approver_emails = list(
        RoomApprover.objects.filter(
            room=occurrence.room,
            is_active=True,
        ).select_related("user").values_list("user__email", flat=True)
    )

    requester_name = occurrence.series.requested_by.get_full_name() or occurrence.series.requested_by.username
    series_title = occurrence.series.title
    freq_display = occurrence.series.get_frequency_display() if occurrence.series.frequency else "Recurring"

    # Count remaining active occurrences
    active_count = occurrence.series.occurrences.exclude(status='cancelled').count()
    total_count = occurrence.series.occurrences.count()
    cancelled_count = total_count - active_count

    subject = f"[Room Booking] Single occurrence cancelled: {occurrence.room.name} on {occurrence.date}"
    message = (
        f"Dear Approver,\n\n"
        f"A single occurrence from a recurring booking series has been CANCELLED by the requester.\n\n"
        f"Cancelled Occurrence:\n"
        f"  Room: {occurrence.room.name} ({occurrence.room.code})\n"
        f"  Date: {occurrence.date}\n"
        f"  Time: {occurrence.start_time} – {occurrence.end_time}\n"
        f"  Title: {series_title}\n"
        f"  Cancelled by: {requester_name}\n\n"
        f"Series Information:\n"
        f"  Frequency: {freq_display}\n"
        f"  Total occurrences: {total_count}\n"
        f"  Active occurrences: {active_count}\n"
        f"  Cancelled occurrences: {cancelled_count}\n\n"
        f"The room is now available for this specific time slot.\n"
        f"The remaining {active_count} occurrences in the series are still active.\n\n"
        f"Thank you."
    )

    _send_email_async(subject, message, approver_emails)


def send_booking_calendar_invite(booking):
    """
    Sends a calendar invite to the room resource, the requester, and all guests.
    """
    # Do not send an invite if the room is not configured for calendar synchronization.
    if not getattr(booking.room, 'calendar_sync_enabled', False):
        return

    # Generate the .ics calendar file content
    ics_bytes = generate_booking_ics(booking)

    # --- Build the list of all recipients ---
    recipients = set()

    # 1. Add the requester
    if booking.requested_by and booking.requested_by.email:
        recipients.add(booking.requested_by.email)

    # 2. Add the room's resource email (if it exists)
    if booking.room.resource_email:
        recipients.add(booking.room.resource_email)

    # 3. Add all guest emails
    if booking.attendee_emails:
        guest_emails = [email.strip() for email in booking.attendee_emails.split(',') if email.strip()]
        recipients.update(guest_emails)

    if not recipients:
        return # No one to send to

    # --- Create and send the email ---
    subject = f"Booking: {booking.title} ({booking.room.name})"
    body = "This email contains a calendar invitation for your meeting."
    from_email = settings.DEFAULT_FROM_EMAIL

    msg = EmailMultiAlternatives(
        subject=subject,
        body=body,
        from_email=from_email,
        to=list(recipients) # Convert set to list for sending
    )

    # Attach the calendar data. The content type is crucial for Outlook/O365.
    msg.attach("invite.ics", ics_bytes, "text/calendar; method=REQUEST")

    # Send the email in a background thread to not block the web request
    threading.Thread(target=lambda: msg.send(fail_silently=True)).start()



def _get_ict_emails_for_booking(booking):
    """
    Collect ICT focal-point emails for a booking.
    Priority order:
      1. Any RoomApprover for this room whose User has role=ict_focal
      2. AgencyAssetRoles.ict_custodian for the requester agency
      3. Any User with role=ict_focal in the same agency (fallback)
    """
    emails = []
    room = booking.room

    # 1. Room-level approvers who are ICT focal
    emails.extend(list(
        RoomApprover.objects.filter(
            room=room,
            is_active=True,
            user__role="ict_focal",
        ).select_related("user").values_list("user__email", flat=True)
    ))

    # 2 & 3. Agency-level ICT roles
    if _HAS_AGENCY_ROLES:
        agency = getattr(booking.requested_by, "agency", None)
        if agency:
            try:
                roles = AgencyAssetRoles.objects.get(agency=agency)
                emails.extend(list(roles.ict_custodian.values_list("email", flat=True)))
            except AgencyAssetRoles.DoesNotExist:
                pass
            emails.extend(list(
                _User.objects.filter(
                    agency=agency, role="ict_focal", is_active=True,
                ).exclude(email="").values_list("email", flat=True)
            ))

    return list(dict.fromkeys(e for e in emails if e))


def notify_ict_support_requested(booking):
    """
    Email ICT focal points when a booking requests ICT support.
    Fires at submission time so ICT can plan ahead.
    """
    ict_emails = _get_ict_emails_for_booking(booking)
    if not ict_emails:
        return

    requester_name = booking.requested_by.get_full_name() or booking.requested_by.username
    support_label = {
        "setup": "Before meeting — Setup / AV configuration",
        "during": "During meeting — Live technical support",
    }.get(getattr(booking, "ict_support", ""), "ICT support requested")

    subject = f"[ICT Support Requested] {booking.room.name} — {booking.date}"
    message = (
        "Dear ICT Team,\n\n"
        "A room booking has been submitted that requires ICT support.\n\n"
        f"  Room:         {booking.room.name} ({booking.room.code})\n"
        f"  Date:         {booking.date}\n"
        f"  Time:         {booking.start_time} – {booking.end_time}\n"
        f"  Title:        {booking.title}\n"
        f"  Requested by: {requester_name}\n"
        f"  ICT Support:  {support_label}\n\n"
        "The booking is currently pending approval. "
        "You may want to reach out to the requester in advance "
        "to clarify any technical requirements.\n\n"
        "Thank you."
    )
    _send_email_async(subject, message, ict_emails)


def notify_ict_support_requested_series(series):
    """Same as notify_ict_support_requested but for a recurring series."""

    class _FakeBooking:
        pass

    fake = _FakeBooking()
    fake.room = series.room
    fake.requested_by = series.requested_by

    ict_emails = _get_ict_emails_for_booking(fake)
    if not ict_emails:
        return

    requester_name = series.requested_by.get_full_name() or series.requested_by.username
    freq_display = series.get_frequency_display() if series.frequency else "One-time"
    support_label = {
        "setup": "Before meeting — Setup / AV configuration",
        "during": "During meeting — Live technical support",
    }.get(getattr(series, "ict_support", ""), "ICT support requested")

    subject = f"[ICT Support Requested] Recurring — {series.room.name}"
    message = (
        "Dear ICT Team,\n\n"
        "A RECURRING room booking series has been submitted that requires ICT support.\n\n"
        f"  Room:          {series.room.name} ({series.room.code})\n"
        f"  Title:         {series.title}\n"
        f"  Frequency:     {freq_display}\n"
        f"  Start Date:    {series.start_date}\n"
        f"  End Date:      {series.end_date or 'No end date'}\n"
        f"  Time:          {series.start_time} – {series.end_time}\n"
        f"  Requested by:  {requester_name}\n"
        f"  ICT Support:   {support_label}\n\n"
        "The series is currently pending approval. "
        "Please coordinate with the requester before the first occurrence.\n\n"
        "Thank you."
    )
    _send_email_async(subject, message, ict_emails)


def room_has_active_approvers(room) -> bool:
    return room.room_approver_links.filter(is_active=True).exists()


def compute_initial_status(room):
    if room.approval_mode == "auto":
        return "approved"

    if room.approval_mode == "mixed":
        if not room.room_approver_links.filter(is_active=True).exists():
            return "approved"
        return "pending"

    return "pending"


def _nth_weekday_of_month(year, month, weekday, n):
    """
    Return the date of the nth occurrence of `weekday` (0=Mon..6=Sun) in year/month.
    n=1 → first, n=2 → second, n=3 → third, n=4 → fourth, n=-1 → last.
    Returns None if that occurrence doesn't exist (e.g. 5th Monday in a short month).
    """
    import calendar as _cal
    if n == -1:
        last_day = _cal.monthrange(year, month)[1]
        d = date_cls(year, month, last_day)
        while d.weekday() != weekday:
            d -= timedelta(days=1)
        return d
    # Find the first occurrence of weekday in the month
    first = date_cls(year, month, 1)
    delta = (weekday - first.weekday()) % 7
    first_occ = first + timedelta(days=delta)
    target = first_occ + timedelta(weeks=n - 1)
    if target.month != month:
        return None  # e.g. "5th Monday" doesn't exist this month
    return target


def _advance_months(d, interval):
    """Advance date d by `interval` months, clamping day to last day of target month."""
    import calendar as _cal
    m = (d.month - 1 + interval)
    year = d.year + (m // 12)
    month = (m % 12) + 1
    last = _cal.monthrange(year, month)[1]
    return date_cls(year, month, min(d.day, last))


def iter_recurrence_dates(
        start_date, end_date, frequency, interval=1, weekdays=None,
        monthly_type="day", monthly_week=None, monthly_weekday=None,
):
    """
    Yield dates in the recurrence series.

    Parameters
    ----------
    weekdays        : list[int] for weekly, e.g. [0, 2, 4]  (Mon=0 … Sun=6)
    monthly_type    : 'day'     → fixed day-of-month (legacy behaviour)
                      'weekday' → nth weekday of month (e.g. last Thursday)
    monthly_week    : 1/2/3/4/-1  (used when monthly_type='weekday')
    monthly_weekday : 0–6         (used when monthly_type='weekday')
    """
    if not frequency:
        yield start_date
        return

    if not end_date:
        # safety default: 1 year max if user didn't set end date
        end_date = start_date + timedelta(days=365)

    if frequency == "daily":
        cur = start_date
        step = timedelta(days=interval)
        while cur <= end_date:
            yield cur
            cur += step

    elif frequency == "weekly":
        weekdays = sorted(weekdays or [])
        if not weekdays:
            weekdays = [start_date.weekday()]  # default: same weekday as start

        # start from the week of start_date
        cur = start_date
        while cur <= end_date:
            # for each week, yield selected weekdays
            week_start = cur - timedelta(days=cur.weekday())  # Monday
            for wd in weekdays:
                d = week_start + timedelta(days=wd)
                if d < start_date:
                    continue
                if d > end_date:
                    continue
                yield d
            cur = week_start + timedelta(days=7 * interval)

    elif frequency == "monthly":
        if monthly_type == "weekday" and monthly_week is not None and monthly_weekday is not None:
            # ---- nth weekday of month (e.g. "last Thursday") ----
            cur_year, cur_month = start_date.year, start_date.month
            while True:
                d = _nth_weekday_of_month(cur_year, cur_month, monthly_weekday, monthly_week)
                if d is not None and d >= start_date and d <= end_date:
                    yield d
                # Advance by `interval` months
                m = (cur_month - 1 + interval)
                cur_year = cur_year + (m // 12)
                cur_month = (m % 12) + 1
                # Stop if the start of the next candidate month is already past end_date
                if date_cls(cur_year, cur_month, 1) > end_date:
                    break
        else:
            # ---- fixed day-of-month (legacy) ----
            cur = start_date
            day = start_date.day
            while cur <= end_date:
                yield cur
                cur = _advance_months(date_cls(cur.year, cur.month, day), interval)

    elif frequency == "yearly":
        cur = start_date
        while cur <= end_date:
            yield cur
            import calendar as _cal
            last = _cal.monthrange(cur.year + interval, cur.month)[1]
            cur = date_cls(cur.year + interval, cur.month, min(cur.day, last))


def find_next_available_slot(room, on_date, duration_minutes, after_time=None, search_days=7):
    """
    Scan from `after_time` onward (in 15-min steps) looking for a free
    window of `duration_minutes` on `on_date`.  If none found that day,
    tries the following days up to `search_days` ahead.

    Returns a dict  { 'start': 'HH:MM', 'end': 'HH:MM', 'label': '...',
                      'date': 'YYYY-MM-DD', 'same_day': bool }
    or None if nothing found.
    """
    duration = timedelta(minutes=max(duration_minutes, 15))
    business_start = time_cls(8, 0)
    business_end = time_cls(20, 0)
    step = timedelta(minutes=15)

    for day_offset in range(search_days):
        check_date = on_date + timedelta(days=day_offset)

        # All approved/pending bookings on this date for the room, sorted
        existing = list(
            RoomBooking.objects.filter(
                room=room,
                date=check_date,
                status__in=('approved', 'pending'),
            ).order_by('start_time').values('start_time', 'end_time')
        )

        # Start scanning from business_start (or after_time on first day)
        if day_offset == 0 and after_time:
            # Round up to nearest 15-min slot after the conflict ends
            candidate_start = datetime.combine(check_date, after_time)
            # Round up to 15 min
            minutes_past = (candidate_start.minute % 15)
            if minutes_past:
                candidate_start += timedelta(minutes=15 - minutes_past)
            candidate_start = candidate_start.replace(second=0, microsecond=0)
        else:
            candidate_start = datetime.combine(check_date, business_start)

        # Scan in 15-min steps
        while True:
            candidate_end = candidate_start + duration

            # Outside business hours
            if candidate_end.time() > business_end:
                break

            # Check against all bookings on this day
            conflict = False
            for bk in existing:
                bk_start = datetime.combine(check_date, bk['start_time'])
                bk_end = datetime.combine(check_date, bk['end_time'])
                # Overlap: candidate_start < bk_end AND candidate_end > bk_start
                if candidate_start < bk_end and candidate_end > bk_start:
                    conflict = True
                    # Jump past this booking
                    candidate_start = bk_end
                    # Round up to 15 min
                    mp = candidate_start.minute % 15
                    if mp:
                        candidate_start += timedelta(minutes=15 - mp)
                    candidate_start = candidate_start.replace(second=0, microsecond=0)
                    break

            if not conflict:
                # Found a slot!
                start_str = candidate_start.strftime('%H:%M')
                end_str = candidate_end.strftime('%H:%M')

                # Human label
                if day_offset == 0:
                    day_label = 'today'
                elif day_offset == 1:
                    day_label = 'tomorrow'
                else:
                    day_label = check_date.strftime('%A, %d %b')  # e.g. "Monday, 19 May"

                label = f"{start_str} – {end_str}  ({day_label})"
                if day_offset > 1:
                    label = f"{check_date.strftime('%a %d %b')}  {start_str} – {end_str}"

                return {
                    'start': start_str,
                    'end': end_str,
                    'date': check_date.strftime('%Y-%m-%d'),
                    'label': label,
                    'same_day': day_offset == 0,
                }
            # If we jumped candidate_start, loop again without stepping
            # (the inner loop already moved candidate_start past the conflict)

        # If we get here: no slot found today → try next day

    return None

def _booking_has_ended(booking):
    """
    Return True if the booking end datetime is already in the past.
    """
    end_dt = datetime.combine(booking.date, booking.end_time)
    end_dt = timezone.make_aware(end_dt, timezone.get_current_timezone())
    return timezone.now() > end_dt


def _booking_confirmed_attendance_count(booking):
    """
    Count only confirmed attendance records.
    """
    return booking.attendance_records.filter(
        status__in=["present", "approved"]
    ).count()


def _booking_start_dt(booking):
    dt = datetime.combine(booking.date, booking.start_time)
    return timezone.make_aware(dt, timezone.get_current_timezone())


def _booking_end_dt(booking):
    dt = datetime.combine(booking.date, booking.end_time)
    return timezone.make_aware(dt, timezone.get_current_timezone())


def _booking_has_started(booking):
    return timezone.now() >= _booking_start_dt(booking)


def _booking_has_ended(booking):
    return timezone.now() > _booking_end_dt(booking)


def _booking_confirmed_attendance_count(booking):
    return booking.attendance_records.filter(
        status__in=["present", "approved"]
    ).count()


def _booking_registered_count(booking):
    return booking.attendees.count()


def _booking_public_link_status(booking):
    """
    Returns a machine code + human label + color class for public page state.
    """
    if not booking.enable_invite_link:
        return {
            "code": "disabled",
            "label": "Public link disabled",
            "badge_class": "secondary",
            "reason": "The booking owner has disabled the public link.",
        }

    if _booking_has_ended(booking):
        return {
            "code": "ended",
            "label": "Meeting ended",
            "badge_class": "dark",
            "reason": "The meeting has already ended.",
        }

    capacity = booking.room.capacity or 0
    if capacity > 0:
        current = (
            _booking_confirmed_attendance_count(booking)
            if booking.enable_attendance
            else _booking_registered_count(booking)
        )
        if current >= capacity:
            return {
                "code": "full",
                "label": "Capacity reached",
                "badge_class": "warning",
                "reason": "The meeting has reached its room capacity.",
            }

    if _booking_has_started(booking):
        return {
            "code": "live",
            "label": "Live now",
            "badge_class": "success",
            "reason": "The meeting is currently in progress.",
        }

    return {
        "code": "open",
        "label": "Open",
        "badge_class": "primary",
        "reason": "The public link is active.",
    }



def _booking_public_link_block_reason(booking):
    status = _booking_public_link_status(booking)
    if status["code"] in {"disabled", "ended", "full"}:
        return status["reason"]
    return None
# ======================= VIEWS =======================


class RoomListView(LoginRequiredMixin, ListView):
    model = Room
    template_name = "accounts/rooms/room_list.html"
    context_object_name = "rooms"

    def get_queryset(self):
        return Room.objects.filter(is_active=True).prefetch_related("amenities")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        import json
        from datetime import datetime as _dt

        user = self.request.user
        today = timezone.localtime().date()
        rooms = ctx["rooms"]

        # ── Stats for the banner ──────────────────────────────────────────
        ctx["total_capacity"] = sum(r.capacity or 0 for r in rooms)
        ctx["available_now"] = sum(1 for r in rooms if getattr(r, "is_available_now", False))
        ctx["bookings_today"] = RoomBooking.objects.filter(
            room__in=rooms, date=today, status="approved"
        ).count()

        # ── Pending approvals badge ───────────────────────────────────────
        ctx["pending_approvals"] = RoomBooking.objects.filter(
            status="pending",
            room__room_approver_links__user=user,
            room__room_approver_links__is_active=True,
            series__isnull=True,
        ).distinct().count()

        # ── Per-room today bookings as JSON for live JS ───────────────────
        # Single query — no N+1
        today_bookings = (
            RoomBooking.objects
            .filter(room__in=rooms, date=today, status="approved")
            .values("room_id", "start_time", "end_time", "title")
        )

        bookings_by_room = {}
        for b in today_bookings:
            rid = str(b["room_id"])  # JS uses String(roomId)
            bookings_by_room.setdefault(rid, []).append({
                "start": b["start_time"].strftime("%H:%M"),
                "end": b["end_time"].strftime("%H:%M"),
                "title": b["title"],
            })

        ctx["bookings_by_room_json"] = json.dumps(bookings_by_room)
        return ctx


class RoomDetailView(LoginRequiredMixin, DetailView):
    model = Room
    template_name = "accounts/rooms/room_detail.html"
    context_object_name = "room"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        room = self.object
        user = self.request.user

        today = timezone.localtime().date()
        future_limit = today + timedelta(days=30)

        # ── Approved upcoming bookings (list view) ──────────────────────
        approved_bookings = list(
            room.bookings.filter(
                date__gte=today,
                date__lte=future_limit,
                status="approved",
            )
            .select_related("requested_by")
            .order_by("date", "start_time")
        )

        from datetime import datetime as _dt
        for b in approved_bookings:
            start = _dt.combine(today, b.start_time)
            end = _dt.combine(today, b.end_time)
            mins = max(int((end - start).total_seconds() / 60), 0)
            b.duration_minutes = mins
            b.duration_class = (
                "heavy" if mins > 180 else
                "medium" if mins > 60 else
                "light"
            )

        ctx["approved_bookings"] = approved_bookings
        ctx["upcoming_bookings"] = approved_bookings

        # ── Timeline: today's approved bookings ─────────────────────────
        TIMELINE_START_HOUR = 7
        timeline_bookings = list(
            room.bookings.filter(date=today, status="approved")
            .select_related("requested_by")
            .order_by("start_time")
        )

        for b in timeline_bookings:
            top = (b.start_time.hour - TIMELINE_START_HOUR) * 60 + b.start_time.minute
            start = _dt.combine(today, b.start_time)
            end = _dt.combine(today, b.end_time)
            mins = max(int((end - start).total_seconds() / 60), 15)
            b.timeline_top = max(top, 0)
            b.timeline_height = mins
            b.duration_minutes = mins

        ctx["timeline_bookings_today"] = timeline_bookings
        ctx["timeline_hours"] = list(range(TIMELINE_START_HOUR, 22))

        # ── Sidebar: my pending bookings ────────────────────────────────
        ctx["my_pending_bookings"] = list(
            room.bookings.filter(requested_by=user, status="pending")
            .order_by("date", "start_time")[:5]
        )

        # ── Is current user an approver for this room? ──────────────────
        ctx["is_approver"] = room.room_approver_links.filter(
            user=user, is_active=True
        ).exists()

        # ── Stats ────────────────────────────────────────────────────────
        ctx["bookings_today"] = room.bookings.filter(
            date=today, status="approved"
        ).count()

        total_mins = sum(b.duration_minutes for b in timeline_bookings)
        ctx["utilization_rate"] = min(round(total_mins / 480 * 100), 100)

        return ctx


class MyRoomBookingsView(LoginRequiredMixin, ListView):
    """
    Enhanced view showing user's bookings with:
    - Recurring series grouped as one item
    - Pagination
    - Search functionality
    - Advanced filters
    """
    template_name = "accounts/rooms/my_bookings.html"
    context_object_name = "items"
    paginate_by = 10  # Number of items per page

    def get_queryset(self):
        user = self.request.user

        # Get all bookings for this user
        bookings_qs = RoomBooking.objects.filter(
            requested_by=user
        ).select_related('room', 'series', 'approved_by')

        # Get all series for this user
        series_qs = RoomBookingSeries.objects.filter(
            requested_by=user
        ).select_related('room', 'approved_by').annotate(
            occurrence_count=Count('occurrences')
        ).prefetch_related(
            Prefetch(
                'occurrences',
                queryset=RoomBooking.objects.order_by('date', 'start_time')
            )
        )

        # Apply filters
        status_filter = self.request.GET.get('status', '')
        when_filter = self.request.GET.get('when', '')
        search_query = self.request.GET.get('search', '')
        date_from = self.request.GET.get('date_from', '')
        date_to = self.request.GET.get('date_to', '')
        room_filter = self.request.GET.get('room', '')

        # Status filter
        if status_filter:
            bookings_qs = bookings_qs.filter(status=status_filter)
            series_qs = series_qs.filter(status=status_filter)

        # Time filter
        today = date.today()
        if when_filter == 'upcoming':
            bookings_qs = bookings_qs.filter(date__gte=today)
            series_qs = series_qs.filter(
                Q(end_date__gte=today) | Q(end_date__isnull=True)
            )
        elif when_filter == 'past':
            bookings_qs = bookings_qs.filter(date__lt=today)
            series_qs = series_qs.filter(end_date__lt=today)

        # Search filter
        if search_query:
            bookings_qs = bookings_qs.filter(
                Q(title__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(room__name__icontains=search_query)
            )
            series_qs = series_qs.filter(
                Q(title__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(room__name__icontains=search_query)
            )

        # Date range filter
        if date_from:
            bookings_qs = bookings_qs.filter(date__gte=date_from)
            series_qs = series_qs.filter(start_date__gte=date_from)

        if date_to:
            bookings_qs = bookings_qs.filter(date__lte=date_to)
            series_qs = series_qs.filter(
                Q(end_date__lte=date_to) | Q(end_date__isnull=True)
            )

        # Room filter
        if room_filter:
            bookings_qs = bookings_qs.filter(room_id=room_filter)
            series_qs = series_qs.filter(room_id=room_filter)

        # Get standalone bookings (not part of any series)
        standalone_bookings = bookings_qs.filter(series__isnull=True)

        # Create combined list with type markers
        items = []

        # Add series
        for series in series_qs:
            items.append({
                'is_series': True,
                'series': series,
                'occurrence_count': series.occurrence_count,
                'occurrences': series.occurrences.all(),
                'sort_date': series.start_date,
                'sort_created': series.created_at,
                'title': series.title,
            })

        # Add standalone bookings
        for booking in standalone_bookings:
            items.append({
                'is_series': False,
                'booking': booking,
                'room': booking.room,
                'approved_by': booking.approved_by,
                'sort_date': booking.date,
                'sort_created': booking.created_at,
                'title': booking.title,
            })

        # Sort
        sort_by = self.request.GET.get('sort', '-created_at')

        if sort_by == 'date':
            items.sort(key=lambda x: x['sort_date'])
        elif sort_by == '-date':
            items.sort(key=lambda x: x['sort_date'], reverse=True)
        elif sort_by == 'title':
            items.sort(key=lambda x: x['title'].lower())
        elif sort_by == '-title':
            items.sort(key=lambda x: x['title'].lower(), reverse=True)
        elif sort_by == 'created_at':
            items.sort(key=lambda x: x['sort_created'])
        else:  # -created_at (default)
            items.sort(key=lambda x: x['sort_created'], reverse=True)

        # Convert to objects that can be used in templates
        class Item:
            def __init__(self, data):
                for key, value in data.items():
                    setattr(self, key, value)

        return [Item(item) for item in items]

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user

        # --- Automatic Survey Logic ---
        now = timezone.now()
        # Find approved bookings that ended in the last 7 days and haven't been surveyed
        unsurveyed_bookings = RoomBooking.objects.filter(
            requested_by=user,
            status='approved',
            date__lt=now.date(),
            survey_sent_at__isnull=True
        ).select_related('room')

        for booking in unsurveyed_bookings:
            # Placeholder for sending email; in a real app, you'd use a proper function
            print(f"Sending survey for booking #{booking.id} for room {booking.room.name}")
            # notify_user_of_survey(booking) # This function would be defined in your email helpers

            # Mark as surveyed
            booking.survey_sent_at = now
            booking.save(update_fields=['survey_sent_at'])

        # Get all bookings and series for stats
        all_bookings = RoomBooking.objects.filter(requested_by=user)
        all_series = RoomBookingSeries.objects.filter(requested_by=user)

        today = date.today()

        # Calculate stats
        ctx['total_bookings'] = all_bookings.count() + all_series.count()
        ctx['pending_count'] = (
                all_bookings.filter(status='pending', series__isnull=True).count() +
                all_series.filter(status='pending').count()
        )
        ctx['approved_count'] = (
                all_bookings.filter(status='approved', series__isnull=True).count() +
                all_series.filter(status='approved').count()
        )
        ctx['upcoming_count'] = (
                all_bookings.filter(
                    date__gte=today,
                    status='approved',
                    series__isnull=True
                ).count() +
                all_series.filter(
                    status='approved'
                ).filter(
                    Q(end_date__gte=today) | Q(end_date__isnull=True)
                ).count()
        )
        ctx['series_count'] = all_series.count()

        # Pass filter values
        ctx['status_filter'] = self.request.GET.get('status', '')
        ctx['when_filter'] = self.request.GET.get('when', '')
        ctx['search_query'] = self.request.GET.get('search', '')
        ctx['date_from'] = self.request.GET.get('date_from', '')
        ctx['date_to'] = self.request.GET.get('date_to', '')
        ctx['room_filter'] = self.request.GET.get('room', '')
        ctx['sort_by'] = self.request.GET.get('sort', '-created_at')

        # Get available rooms for filter dropdown
        ctx['available_rooms'] = Room.objects.filter(is_active=True).order_by('name')

        return ctx


class RoomBookingCreateView(LoginRequiredMixin, CreateView):
    model = RoomBooking
    form_class = RoomBookingForm
    template_name = "accounts/rooms/booking_form.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        room_id = self.request.GET.get("room")
        if room_id:
            kwargs["room"] = get_object_or_404(Room, pk=room_id)
        return kwargs

    def get_success_url(self):
        return reverse("accounts:my_bookings")

    def get_form(self, form_class=None):
        """
        Pass the selected room into the form when present.
        """
        if form_class is None:
            form_class = self.get_form_class()

        form_kwargs = self.get_form_kwargs()
        room_id = self.request.GET.get("room")

        if room_id:
            try:
                room = Room.objects.get(pk=room_id)
                return form_class(room=room, **form_kwargs)
            except (Room.DoesNotExist, ValueError, TypeError):
                pass

        return form_class(**form_kwargs)

    def form_valid(self, form):
        """
        IMPORTANT:
        Django calls form_valid(), not form_valid_with_conflict_guard().
        """
        return self.form_valid_with_conflict_guard(form)

    def form_valid_with_conflict_guard(self, form):
        from django.core.exceptions import ValidationError as DjangoValidationError

        user = self.request.user
        room = form.cleaned_data.get("room") or get_object_or_404(
            Room, pk=self.request.GET.get("room")
        )
        frequency = form.cleaned_data.get("frequency")
        status = compute_initial_status(room)

        attendee_emails = form.cleaned_data.get("attendee_emails", "") or ""
        virtual_meeting_link = form.cleaned_data.get("virtual_meeting_link", "") or ""
        selected_amenities = form.cleaned_data.get("selected_amenities")
        ict_support = form.cleaned_data.get("ict_support", "none") or "none"

        def make_conflict_context(room, booking_date, start_time, end_time):
            existing = RoomBooking.objects.filter(
                room=room,
                date=booking_date,
                status__in=("approved", "pending"),
                start_time__lt=end_time,
                end_time__gt=start_time,
            ).order_by("start_time")

            ci = existing.first()

            req_start = datetime.combine(booking_date, start_time)
            req_end = datetime.combine(booking_date, end_time)
            req_duration = int((req_end - req_start).total_seconds() // 60)

            after = ci.end_time if ci else None
            next_slot = find_next_available_slot(
                room=room,
                on_date=booking_date,
                duration_minutes=req_duration,
                after_time=after,
            )
            return ci, next_slot

        # =========================
        # RECURRING BOOKING PATH
        # =========================
        if frequency:
            until = form.cleaned_data.get("until")
            interval = form.cleaned_data.get("interval", 1)

            weekdays_raw = form.cleaned_data.get("weekdays", [])
            weekdays = [int(x) for x in weekdays_raw] if weekdays_raw else []

            monthly_type = form.cleaned_data.get("monthly_type", "day") or "day"
            monthly_week_raw = form.cleaned_data.get("monthly_week")
            monthly_weekday_raw = form.cleaned_data.get("monthly_weekday")

            monthly_week = (
                int(monthly_week_raw)
                if monthly_week_raw not in (None, "")
                else None
            )
            monthly_weekday = (
                int(monthly_weekday_raw)
                if monthly_weekday_raw not in (None, "")
                else None
            )

            try:
                with transaction.atomic():
                    series = RoomBookingSeries.objects.create(
                        room=room,
                        requested_by=user,
                        title=form.cleaned_data["title"],
                        description=form.cleaned_data.get("description", ""),
                        start_date=form.cleaned_data["date"],
                        end_date=until,
                        start_time=form.cleaned_data["start_time"],
                        end_time=form.cleaned_data["end_time"],
                        frequency=frequency,
                        interval=interval,
                        weekdays_csv=",".join(str(x) for x in (weekdays or [])),
                        monthly_type=monthly_type,
                        monthly_week=monthly_week,
                        monthly_weekday=monthly_weekday,
                        status=status,
                        ict_support=ict_support,
                        attendee_emails=attendee_emails,
                        virtual_meeting_link=virtual_meeting_link,
                    )

                    created = 0
                    first_booking = None

                    for d in iter_recurrence_dates(
                        start_date=series.start_date,
                        end_date=series.end_date,
                        frequency=frequency,
                        interval=interval,
                        weekdays=weekdays,
                        monthly_type=monthly_type,
                        monthly_week=monthly_week,
                        monthly_weekday=monthly_weekday,
                    ):
                        booking = RoomBooking(
                            series=series,
                            room=room,
                            title=series.title,
                            description=series.description,
                            date=d,
                            start_time=series.start_time,
                            end_time=series.end_time,
                            status=status,
                            requested_by=user,
                            ict_support=ict_support,
                            attendee_emails=attendee_emails,
                            virtual_meeting_link=virtual_meeting_link,
                            enable_attendance=form.cleaned_data.get("enable_attendance", False),
                            enable_invite_link=form.cleaned_data.get("enable_invite_link", False),
                        )

                        booking.full_clean()
                        booking.save()

                        if selected_amenities:
                            booking.selected_amenities.set(selected_amenities)
                            booking.requested_amenities.set(selected_amenities)  # mirror for detail page display

                        if booking.enable_invite_link and not booking.registration_code:
                            import uuid
                            booking.registration_code = uuid.uuid4()
                            booking.save(update_fields=["registration_code"])

                        if created == 0:
                            first_booking = booking
                        created += 1

                    if status == "pending":
                        notify_approvers_new_series(series)
                        notify_requester_series_submitted(series)
                    else:
                        if getattr(room, "auto_approve_notify_approvers", False):
                            notify_approvers_new_series(series)
                        notify_requester_series_approved(series)
                        if first_booking:
                            send_booking_calendar_invite(first_booking)

                    if ict_support != "none":
                        notify_ict_support_requested_series(series)

                    messages.success(
                        self.request,
                        f"Recurring booking created ({created} occurrences)"
                        + (" and auto-approved." if status == "approved" else " — awaiting approval.")
                    )
                    return redirect("accounts:room_detail", pk=room.pk)

            except DjangoValidationError as exc:
                ci, next_slot = make_conflict_context(
                    room,
                    form.cleaned_data["date"],
                    form.cleaned_data["start_time"],
                    form.cleaned_data["end_time"],
                )
                form.add_error(None, str(exc))
                return self.render_to_response(
                    self.get_context_data(
                        form=form,
                        conflict_info=ci,
                        next_available=next_slot,
                    )
                )

        # =========================
        # SINGLE BOOKING PATH
        # =========================
        form.instance.requested_by = user
        form.instance.room = room
        form.instance.ict_support = ict_support
        form.instance.status = status

        try:
            form.instance.full_clean()
        except DjangoValidationError as exc:
            ci, next_slot = make_conflict_context(
                room,
                form.cleaned_data["date"],
                form.cleaned_data["start_time"],
                form.cleaned_data["end_time"],
            )
            form.add_error(None, str(exc))
            return self.render_to_response(
                self.get_context_data(
                    form=form,
                    conflict_info=ci,
                    next_available=next_slot,
                )
            )

        response = super().form_valid(form)

        if selected_amenities:
            self.object.selected_amenities.set(selected_amenities)
            self.object.requested_amenities.set(selected_amenities)  # mirror for detail page display

        if self.object.enable_invite_link and not self.object.registration_code:
            import uuid
            self.object.registration_code = uuid.uuid4()
            self.object.save(update_fields=["registration_code"])

        if status == "pending":
            notify_approvers_new_booking(self.object)
            notify_requester_booking_submitted(self.object)
            messages.success(
                self.request,
                "Booking request submitted and awaiting approval."
            )
        else:
            self.object.status = "approved"
            self.object.approved_by = None
            self.object.approved_at = timezone.now()
            self.object.save(update_fields=["status", "approved_by", "approved_at"])

            if getattr(room, "auto_approve_notify_approvers", False):
                notify_approvers_new_booking(self.object)

            notify_requester_booking_approved(self.request, self.object)
            send_booking_calendar_invite(self.object)

            messages.success(
                self.request,
                "Booking auto-approved. Calendar invites sent."
            )

        if ict_support != "none":
            notify_ict_support_requested(self.object)

        return response


class MyRoomApprovalsView(LoginRequiredMixin, ListView):
    """
    Show PENDING bookings AND series where the current user is an active approver.
    """
    template_name = "accounts/rooms/approvals_list.html"
    context_object_name = "items"

    def get_queryset(self):
        """
        Return a combined list of pending individual bookings and pending series.
        We'll mark each with a type field for template rendering.
        """
        user = self.request.user

        # Get pending individual bookings (not part of a series, or series already approved)
        individual_bookings = list(
            RoomBooking.objects.filter(
                status="pending",
                room__room_approver_links__user=user,
                room__room_approver_links__is_active=True,
                series__isnull=True,  # Only non-series bookings
            )
            .select_related("room", "requested_by")
            .order_by("date", "start_time")
            .distinct()
        )

        # Get pending series
        pending_series = list(
            RoomBookingSeries.objects.filter(
                status="pending",
                room__room_approver_links__user=user,
                room__room_approver_links__is_active=True,
            )
            .select_related("room", "requested_by")
            .annotate(occurrence_count=Count("occurrences"))
            .order_by("start_date", "start_time")
            .distinct()
        )

        # Tag each item with its type
        for booking in individual_bookings:
            booking.item_type = "booking"

        for series in pending_series:
            series.item_type = "series"

        # Combine and sort
        all_items = individual_bookings + pending_series

        # Sort by date (use start_date for series, date for bookings)
        all_items.sort(key=lambda x: (
            x.start_date if hasattr(x, 'start_date') else x.date,
            x.start_time
        ))

        return all_items

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        items = ctx['items']

        # Calculate stats
        total_count = len(items)
        series_count = sum(1 for i in items if hasattr(i, 'item_type') and i.item_type == 'series')
        booking_count = total_count - series_count

        today = timezone.localtime().date()
        urgent_count = sum(1 for i in items if (
                (hasattr(i, 'date') and i.date <= today + timedelta(days=2)) or
                (hasattr(i, 'start_date') and i.start_date <= today + timedelta(days=2))
        ))

        ctx.update({
            'total_count': total_count,
            'series_count': series_count,
            'booking_count': booking_count,
            'urgent_count': urgent_count,
            'today_date': today,  # Add this for template usage
            'today_count': sum(1 for i in items if (
                    (hasattr(i, 'date') and i.date == today) or
                    (hasattr(i, 'start_date') and i.start_date == today)
            )),
            'this_week_count': sum(1 for i in items if (
                    (hasattr(i, 'date') and i.date <= today + timedelta(days=7)) or
                    (hasattr(i, 'start_date') and i.start_date <= today + timedelta(days=7))
            )),
        })

        return ctx


@login_required
def room_booking_approve_view(request, pk):
    """
    Approve or reject an individual booking and confirm available amenities.
    """
    booking = get_object_or_404(RoomBooking, pk=pk)
    room = booking.room

    # Permission check: User must be a superuser or an active approver for this room.
    is_approver = request.user.is_superuser or RoomApprover.objects.filter(
        room=room,
        user=request.user,
        is_active=True,
    ).exists()

    if not is_approver:
        messages.error(request, "You are not an approver for this room.")
        return redirect("accounts:room_detail", pk=room.pk)

    if booking.status != "pending":
        messages.info(request, "This booking has already been processed.")
        return redirect("accounts:room_approvals")

    if request.method == "POST":
        # We pass the booking instance to the form
        form = RoomBookingApprovalForm(request.POST, instance=booking)

        # The form is valid if the submitted amenity IDs are valid
        if form.is_valid():
            action = request.POST.get('action')  # 'approve' or 'reject' from button name

            if action == "approve":
                # Save the form to update the `approved_amenities` on the booking instance
                booking = form.save()

                ensure_booking_registration_code(booking)

                # Finalize the approval
                booking.approve(request.user)

                # Send notifications
                notify_requester_booking_approved(request, booking)
                send_booking_calendar_invite(booking)

                messages.success(request, "Booking approved and calendar invite sent.")
                return redirect("accounts:room_approvals")

            elif action == "reject":
                rejection_reason = form.cleaned_data.get('rejection_reason', '').strip()
                if not rejection_reason:
                    messages.error(request, "A reason is required to reject a booking.")
                    # Re-render the page with the error
                    return render(request, "accounts/rooms/booking_approve.html", {
                        "booking": booking, "room": room, "form": form
                    })

                booking.reject(request.user, reason=rejection_reason)
                notify_requester_booking_rejected(booking)
                messages.warning(request, "Booking has been rejected.")
                return redirect("accounts:room_approvals")

            else:
                messages.error(request, "Invalid action specified.")

    else:
        # For a GET request, initialize the form with the booking instance.
        # This will pre-populate the 'approved_amenities' checklist.
        form = RoomBookingApprovalForm(instance=booking)

    return render(request, "accounts/rooms/booking_approve.html", {
        "booking": booking,
        "room": room,
        "form": form,
    })


@login_required
def room_series_approve_view(request, pk):
    """Approve/reject an entire booking series. This view remains unchanged."""
    series = get_object_or_404(
        RoomBookingSeries.objects.select_related('room', 'requested_by')
        .annotate(occurrence_count=Count('occurrences')),
        pk=pk
    )
    room = series.room

    is_approver = request.user.is_superuser or RoomApprover.objects.filter(
        room=room,
        user=request.user,
        is_active=True,
    ).exists()

    if not is_approver:
        messages.error(request, "You are not an approver for this room.")
        return redirect("accounts:room_detail", pk=room.pk)

    if series.status != "pending":
        messages.info(request, "This series has already been processed.")
        return redirect("accounts:room_approvals")

    if request.method == "POST":
        form = RoomSeriesApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data["action"]
            reason = form.cleaned_data["reason"]

            if action == "approve":
                series.approve(request.user)
                notify_requester_series_approved(series)

                ensure_booking_registration_code(booking)

                # Send invite for the first occurrence, which contains the recurrence rule
                first_booking = series.occurrences.order_by('date', 'start_time').first()
                if first_booking:
                    send_booking_calendar_invite(first_booking)

                messages.success(request,
                                 f"Series approved! All {series.occurrences.count()} occurrences have been approved.")
            else:
                if not reason.strip():
                    messages.error(request, "Please provide a reason for rejection.")
                    return render(request, "accounts/rooms/series_approve.html",
                                  {"series": series, "room": room, "form": form})

                series.reject(request.user, reason=reason)
                notify_requester_series_rejected(series)
                messages.warning(request,
                                 f"Series rejected. All {series.occurrences.count()} occurrences have been rejected.")

            return redirect("accounts:room_approvals")
    else:
        form = RoomSeriesApprovalForm()

    sample_occurrences = series.occurrences.all()[:5]

    return render(request, "accounts/rooms/series_approve.html", {
        "series": series,
        "room": room,
        "form": form,
        "sample_occurrences": sample_occurrences,
    })


@login_required
@login_required
def series_detail_view(request, pk):
    """
    Detail page for a recurring booking series.
    Shows a month-grouped timeline of all occurrences, series metadata,
    completion stats, and per-occurrence cancel/detail actions.
    Only the series owner can view this page.
    """
    series = get_object_or_404(
        RoomBookingSeries.objects.select_related(
            'room', 'requested_by', 'approved_by'
        ),
        pk=pk,
        requested_by=request.user,
    )
    occ_list = list(
        series.occurrences
        .select_related('room')
        .order_by('date', 'start_time')
    )

    today = date.today()
    total     = len(occ_list)
    approved  = sum(1 for o in occ_list if o.status == 'approved')
    pending   = sum(1 for o in occ_list if o.status == 'pending')
    cancelled = sum(1 for o in occ_list if o.status == 'cancelled')
    rejected  = sum(1 for o in occ_list if o.status == 'rejected')
    upcoming  = sum(1 for o in occ_list if o.date >= today and o.status in ('approved', 'pending'))
    past      = sum(1 for o in occ_list if o.date < today)
    next_occ  = next((o for o in occ_list if o.date >= today and o.status in ('approved', 'pending')), None)

    # Group occurrences by month label so template can render month separators cleanly
    from itertools import groupby
    def month_key(o):
        return o.date.strftime('%B %Y')

    month_groups = [
        {'month': month, 'occurrences': list(occs)}
        for month, occs in groupby(occ_list, key=month_key)
    ]

    # Parse invited guest emails into a clean list
    invited_emails = [
        e.strip() for e in (series.attendee_emails or '').split(',') if e.strip()
    ]

    return render(request, 'accounts/rooms/series_detail.html', {
        'series':          series,
        'month_groups':    month_groups,
        'today':           today,
        'next_occurrence': next_occ,
        'invited_emails':  invited_emails,
        'stats': {
            'total':    total,
            'approved': approved,
            'pending':  pending,
            'cancelled':cancelled,
            'rejected': rejected,
            'upcoming': upcoming,
            'past':     past,
        },
    })


def booking_detail_view(request, pk):
    booking = get_object_or_404(
        RoomBooking.objects.select_related('room', 'requested_by', 'requested_by__agency')
        .prefetch_related(
            'attendees',
            'approved_amenities',
            'requested_amenities',
            'selected_amenities',
            'attendance_records',
        ),
        pk=pk,
    )

    # Only generate a registration_code when the public link feature is actually enabled
    if booking.enable_invite_link and not booking.registration_code:
        import uuid as _uuid
        booking.registration_code = _uuid.uuid4()
        booking.save(update_fields=['registration_code'])

    if booking.registration_code:
        registration_link = request.build_absolute_uri(
            reverse('accounts:meeting_registration', args=[booking.registration_code])
        )
        qr_download_link = reverse('accounts:meeting_qr_code_download', args=[booking.registration_code])
    else:
        registration_link = None
        qr_download_link = None

    all_records = booking.attendance_records.all()
    pending_walkins = [r for r in all_records if r.status == 'pending_approval']
    confirmed_count = sum(1 for r in all_records if r.status in ('present', 'approved'))
    pending_count = len(pending_walkins)

    if not booking.enable_attendance:
        attendance_records = booking.attendees.all()
        confirmed_count = attendance_records.count()
    else:
        attendance_records = all_records

    public_status = _booking_public_link_status(booking)
    meeting_started = _booking_has_started(booking)
    meeting_ended = _booking_has_ended(booking)

    # ── Invite list: combine invited emails + registered attendees ──────────
    # Parse the comma-separated invited emails into a set for fast lookup
    invited_email_set = {
        e.strip().lower()
        for e in (booking.attendee_emails or '').split(',')
        if e.strip()
    }

    # Build the registered list with their attendance status merged in
    registered_attendees = list(booking.attendees.all().order_by('registered_at'))

    # Build a map: email → attendance record (for cross-referencing check-in status)
    attendance_map = {
        r.email.lower(): r for r in all_records
    }

    # Build a unified invite list entry per unique email
    # Each entry: { email, name, organization, gender, phone,
    #               source (invited/registered/both), registered_at,
    #               attendance_record (or None), attendee_obj (or None) }
    invite_list = []
    seen_emails = set()

    # Start with registered attendees (they have the most info)
    for att in registered_attendees:
        email = att.email.lower()
        seen_emails.add(email)
        att_record = attendance_map.get(email)
        invite_list.append({
            'email': att.email,
            'name': att.name,
            'organization': getattr(att, 'organization', ''),
            'gender': getattr(att, 'gender', ''),
            'phone': getattr(att, 'phone', ''),
            'source': 'both' if email in invited_email_set else 'registered',
            'registered_at': att.registered_at,
            'attendance_record': att_record,
            'attendee_obj': att,
            'attendee_pk': att.pk,
            'is_accepted': getattr(att, 'is_accepted', True),
        })

    # Add pure invite-only entries (emailed but haven't registered yet)
    for email in sorted(invited_email_set):
        if email not in seen_emails:
            att_record = attendance_map.get(email)
            invite_list.append({
                'email': email,
                'name': '',
                'organization': '',
                'gender': '',
                'phone': '',
                'source': 'invited',
                'registered_at': None,
                'attendance_record': att_record,
                'attendee_obj': None,
                'attendee_pk': None,
                'is_accepted': True,
            })

    # Pending registrations (not auto-accepted): registered but not yet accepted
    pending_registrations = [
        e for e in invite_list
        if e['attendee_obj'] is not None and not e.get('is_accepted', True)
    ]

    # Agenda document QR URL
    agenda_qr_url = None
    if booking.agenda_document:
        agenda_qr_url = reverse('accounts:booking_agenda_qr', args=[booking.pk])

    # ── Gate attention flags for meeting-linked visitor access requests ───────
    # Find any GroupMembers linked to this booking's visitor access requests
    # that have been flagged by gate security.
    gate_flagged_members = []
    try:
        from visitors.models import Visitor as VisitorModel, GroupMember as GroupMemberModel
        linked_visitors = VisitorModel.objects.filter(linked_booking=booking)
        for v in linked_visitors:
            for m in v.group_members.filter(gate_attention='needs_attention'):
                gate_flagged_members.append({
                    'member_pk': m.pk,
                    'name': m.full_name,
                    'email': m.email,
                    'note': m.gate_attention_note,
                    'raised_at': m.gate_attention_raised_at.strftime('%H:%M') if m.gate_attention_raised_at else '',
                    'visitor_pk': v.pk,
                    'attendee_pk': m.meeting_attendee_id,
                })
    except Exception:
        pass

    # Enrich invite_list entries with gate_flagged flag
    # Match by attendee_pk (meeting_attendee_id on GroupMember)
    flagged_attendee_pks = {f['attendee_pk'] for f in gate_flagged_members if f['attendee_pk']}
    flagged_emails = {f['email'].lower() for f in gate_flagged_members if f['email']}
    for entry in invite_list:
        entry_email = (entry.get('email') or '').lower()
        entry_pk = entry.get('attendee_pk')
        is_flagged = (entry_pk and entry_pk in flagged_attendee_pks) or (entry_email in flagged_emails)
        entry['gate_flagged'] = is_flagged
        if is_flagged:
            # Find the matching flag to get note
            for f in gate_flagged_members:
                if (f['attendee_pk'] == entry_pk) or (f['email'].lower() == entry_email):
                    entry['gate_flag_note'] = f['note']
                    break

    return render(request, 'accounts/rooms/booking_detail.html', {
        'booking': booking,
        'registration_link': registration_link,
        'qr_download_link': qr_download_link,
        'attendance_records': attendance_records,
        'pending_walkins': pending_walkins,
        'confirmed_count': confirmed_count,
        'pending_count': pending_count,
        'public_status': public_status,
        'meeting_started': meeting_started,
        'meeting_ended': meeting_ended,
        'agenda_qr_url': agenda_qr_url,
        'invite_list': invite_list,
        'invited_count': len(invited_email_set),
        'registered_count': len(registered_attendees),
        'pending_registrations': pending_registrations,
        'gate_flagged_members': gate_flagged_members,
    })


@login_required
@require_GET
def room_detail_api(request, pk):
    """
    Returns JSON with room approval_mode, amenities, and capacity.
    Used by the booking form to show available amenities for manual-approval rooms.
    """
    room = get_object_or_404(Room, pk=pk, is_active=True)
    amenities = [
        {
            'id': a.id,
            'name': a.name,
            'icon_class': a.icon_class or 'bi bi-check-circle',
            'description': a.description,
        }
        for a in room.amenities.filter(is_active=True)
    ]
    return JsonResponse({
        'id': room.pk,
        'name': room.name,
        'approval_mode': room.approval_mode,
        'capacity': room.capacity,
        'amenities': amenities,
    })


@require_GET
def attendance_checkin_lookup(request, registration_code):
    booking = get_object_or_404(RoomBooking, registration_code=registration_code)

    block_reason = _booking_public_link_block_reason(booking)
    if block_reason:
        return JsonResponse({
            'closed': True,
            'message': block_reason,
        }, status=410)

    email = (request.GET.get('email') or '').strip().lower()

    if not email:
        return JsonResponse({'error': 'Email required'}, status=400)

    already = AttendanceRecord.objects.filter(
        booking=booking,
        email__iexact=email
    ).exists()
    if already:
        return JsonResponse({'already_checked_in': True})

    invited_emails = [
        e.strip().lower()
        for e in (booking.attendee_emails or '').split(',')
        if e.strip()
    ]
    is_invited = email in invited_emails

    is_registered = MeetingAttendee.objects.filter(
        booking=booking,
        email__iexact=email
    ).exists()

    return JsonResponse({
        'already_checked_in': False,
        'is_invited': is_invited,
        'is_registered': is_registered,
    })


@login_required
@require_GET
def booking_attendee_count_api(request, pk):
    """
    JSON endpoint returning live attendee count for a booking.
    Called every 30 seconds by the booking_detail template for auto-refresh.
    """
    booking = get_object_or_404(RoomBooking, pk=pk)

    if booking.enable_attendance:
        confirmed_count = booking.attendance_records.filter(
            status__in=["present", "approved"]
        ).count()

        pending_count = booking.attendance_records.filter(
            status="pending_approval"
        ).count()

        return JsonResponse({
            "count": confirmed_count,
            "pending_count": pending_count,
            "booking_id": pk,
            "capacity": booking.room.capacity,
            "mode": "attendance",
        })

    registered_count = booking.attendees.count()

    return JsonResponse({
        "count": registered_count,
        "pending_count": 0,
        "booking_id": pk,
        "capacity": booking.room.capacity,
        "mode": "registration",
    })


def meeting_registration_view(request, registration_code):
    booking = get_object_or_404(
        RoomBooking.objects.select_related('room', 'requested_by', 'requested_by__agency'),
        registration_code=registration_code
    )

    block_reason = _booking_public_link_block_reason(booking)
    if block_reason:
        return render(
            request,
            'accounts/rooms/meeting_registration_closed.html',
            {
                'booking': booking,
                'block_reason': block_reason,
                'public_status': _booking_public_link_status(booking),
            },
            status=410
        )

    if request.method == 'POST':
        form = MeetingAttendeeForm(request.POST)

        if form.is_valid():
            # Re-check on submit in case meeting became blocked while page was open
            block_reason = _booking_public_link_block_reason(booking)
            if block_reason:
                return render(
                    request,
                    'accounts/rooms/meeting_registration_closed.html',
                    {
                        'booking': booking,
                        'block_reason': block_reason,
                        'public_status': _booking_public_link_status(booking),
                    },
                    status=410
                )

            mode = request.POST.get('mode', 'register')
            name = form.cleaned_data['name']
            email = form.cleaned_data['email'].strip().lower()
            organization = form.cleaned_data.get('organization', '')

            if mode == 'attendance' and booking.enable_attendance:
                if AttendanceRecord.objects.filter(booking=booking, email__iexact=email).exists():
                    messages.warning(request, "You have already checked in to this meeting.")
                    return render(
                        request,
                        'accounts/rooms/meeting_registration.html',
                        {
                            'form': MeetingAttendeeForm(),
                            'booking': booking,
                            'public_status': _booking_public_link_status(booking),
                        }
                    )

                invited_emails = [
                    e.strip().lower()
                    for e in (booking.attendee_emails or '').split(',')
                    if e.strip()
                ]
                is_invited = email in invited_emails
                is_registered = MeetingAttendee.objects.filter(
                    booking=booking,
                    email__iexact=email
                ).exists()
                is_auto = is_invited or is_registered
                status = 'present' if is_auto else 'pending_approval'

                AttendanceRecord.objects.create(
                    booking=booking,
                    name=name,
                    email=email,
                    organization=organization,
                    status=status,
                    was_invited=is_invited,
                    was_preregistered=is_registered,
                )

                if is_auto:
                    messages.success(
                        request,
                        f"Welcome, {name}! You've been marked as present. See you inside!"
                    )
                else:
                    messages.warning(
                        request,
                        f"Thank you, {name}. Your check-in request has been sent to the meeting host for approval."
                    )

                return redirect('accounts:meeting_registration_success')

            attendee = form.save(commit=False)
            attendee.booking = booking
            # Save gender and phone if the model supports them (safe before migration)
            gender = request.POST.get('gender', '').strip()
            phone  = request.POST.get('phone', '').strip()
            if hasattr(attendee, 'gender'):
                attendee.gender = gender
            if hasattr(attendee, 'phone'):
                attendee.phone = phone
            # Set is_accepted based on the booking's auto-accept setting
            if hasattr(attendee, 'is_accepted'):
                attendee.is_accepted = getattr(booking, 'auto_accept_registration', True)
            try:
                attendee.save()
                notify_attendee_of_registration(attendee)
                if getattr(booking, 'auto_accept_registration', False):
                    messages.success(
                        request,
                        "You are registered and your spot is confirmed!"
                    )
                else:
                    messages.success(
                        request,
                        "Thank you for registering! A confirmation email has been sent."
                    )
                return redirect('accounts:meeting_registration_success')
            except Exception:
                messages.error(
                    request,
                    "The email address you entered has already been registered for this meeting."
                )
    else:
        form = MeetingAttendeeForm()

    return render(
        request,
        'accounts/rooms/meeting_registration.html',
        {
            'form': form,
            'booking': booking,
            'public_status': _booking_public_link_status(booking),
        }
    )

@login_required
@require_POST
def walkin_decision_view(request, pk, action):
    """
    POST /accounts/attendance/<pk>/approve/  or  /accounts/attendance/<pk>/reject/
    Only the booking requester (or an approver) can call this.
    """
    record = get_object_or_404(AttendanceRecord, pk=pk)
    booking = record.booking

    # Permission: requester or room approver
    is_owner = booking.requested_by == request.user
    is_room_approver = RoomApprover.objects.filter(
        room=booking.room, user=request.user, is_active=True
    ).exists()

    if not (is_owner or is_room_approver or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

    if record.status != 'pending_approval':
        return JsonResponse({'success': False, 'message': 'Record is not pending.'}, status=400)

    from django.utils import timezone as tz
    if action == 'approve':
        record.status = 'approved'
        record.decided_by = request.user
        record.decided_at = tz.now()
        record.save(update_fields=['status', 'decided_by', 'decided_at'])
        return JsonResponse({'success': True, 'status': 'approved'})

    elif action == 'reject':
        record.status = 'rejected'
        record.decided_by = request.user
        record.decided_at = tz.now()
        record.save(update_fields=['status', 'decided_by', 'decided_at'])
        return JsonResponse({'success': True, 'status': 'rejected'})

    return JsonResponse({'success': False, 'message': 'Invalid action.'}, status=400)


def attendance_page_view(request, registration_code):
    """
    Public attendance check-in page — opened when a user scans the meeting QR code.

    Flow
    ----
    Step 1  (GET / POST step=email):
        User enters their email address.
        * Already checked in  →  show status message.
        * Invited or pre-registered  →  show confirm form (name pre-filled).
        * Unknown walk-in  →  show full form (name, org, gender, phone).

    Step 2  (POST step=checkin):
        Create AttendanceRecord.
        * Invited / pre-registered  →  status = 'present'  (auto-confirmed).
        * Walk-in  →  status = 'pending_approval'  (host must approve).
    """
    booking = get_object_or_404(
        RoomBooking.objects.select_related('room', 'requested_by'),
        registration_code=registration_code,
    )

    # If attendance tracking is off, fall back to the standard registration page.
    if not booking.enable_attendance:
        return redirect('accounts:meeting_registration', registration_code=registration_code)

    block_reason = _booking_public_link_block_reason(booking)
    if block_reason:
        return render(
            request,
            'accounts/rooms/meeting_registration_closed.html',
            {
                'booking': booking,
                'block_reason': block_reason,
                'public_status': _booking_public_link_status(booking),
            },
            status=410,
        )

    public_status = _booking_public_link_status(booking)
    step = request.POST.get('step', 'email') if request.method == 'POST' else 'email'

    # ── STEP 1 : email lookup ────────────────────────────────────────────────
    if request.method == 'POST' and step == 'email':
        email = (request.POST.get('email') or '').strip().lower()
        if not email:
            return render(request, 'accounts/rooms/attendance_page.html', {
                'booking': booking,
                'public_status': public_status,
                'step': 'email',
                'error': 'Please enter a valid email address.',
            })

        # Already checked in?
        existing = AttendanceRecord.objects.filter(
            booking=booking, email__iexact=email
        ).first()
        if existing:
            return render(request, 'accounts/rooms/attendance_page.html', {
                'booking': booking,
                'public_status': public_status,
                'step': 'already_in',
                'record': existing,
            })

        invited_emails = [
            e.strip().lower()
            for e in (booking.attendee_emails or '').split(',')
            if e.strip()
        ]
        is_invited    = email in invited_emails
        is_registered = MeetingAttendee.objects.filter(
            booking=booking, email__iexact=email
        ).exists()

        if is_invited or is_registered:
            prefill_name = ''
            if is_registered:
                reg_att = MeetingAttendee.objects.filter(
                    booking=booking, email__iexact=email
                ).first()
                prefill_name = reg_att.name if reg_att else ''
            return render(request, 'accounts/rooms/attendance_page.html', {
                'booking': booking,
                'public_status': public_status,
                'step': 'confirm',
                'email': email,
                'prefill_name': prefill_name,
                'is_invited': is_invited,
                'is_registered': is_registered,
            })

        # Unknown walk-in — collect full details
        return render(request, 'accounts/rooms/attendance_page.html', {
            'booking': booking,
            'public_status': public_status,
            'step': 'walkin_form',
            'email': email,
        })

    # ── STEP 2 : submit check-in ─────────────────────────────────────────────
    if request.method == 'POST' and step == 'checkin':
        email        = (request.POST.get('email')        or '').strip().lower()
        name         = (request.POST.get('name')         or '').strip()
        organization = (request.POST.get('organization') or '').strip()
        gender       = (request.POST.get('gender')       or '').strip()
        phone        = (request.POST.get('phone')        or '').strip()

        if not email or not name:
            return render(request, 'accounts/rooms/attendance_page.html', {
                'booking': booking,
                'public_status': public_status,
                'step': 'email',
                'error': 'Missing required fields. Please try again.',
            })

        # Race-condition guard
        existing = AttendanceRecord.objects.filter(
            booking=booking, email__iexact=email
        ).first()
        if existing:
            return render(request, 'accounts/rooms/attendance_page.html', {
                'booking': booking,
                'public_status': public_status,
                'step': 'already_in',
                'record': existing,
            })

        invited_emails = [
            e.strip().lower()
            for e in (booking.attendee_emails or '').split(',')
            if e.strip()
        ]
        is_invited    = email in invited_emails
        is_registered = MeetingAttendee.objects.filter(
            booking=booking, email__iexact=email
        ).exists()
        is_auto = is_invited or is_registered

        record = AttendanceRecord(
            booking=booking,
            name=name,
            email=email,
            organization=organization,
            status='present' if is_auto else 'pending_approval',
            was_invited=is_invited,
            was_preregistered=is_registered,
        )
        # Store gender / phone only if the model has them (avoids errors before migration)
        if hasattr(AttendanceRecord, 'gender'):
            record.gender = gender
        if hasattr(AttendanceRecord, 'phone'):
            record.phone = phone
        record.save()

        return render(request, 'accounts/rooms/attendance_page.html', {
            'booking': booking,
            'public_status': public_status,
            'step': 'success',
            'name': name,
            'is_auto': is_auto,
        })

    # ── Default GET: show email input ────────────────────────────────────────
    return render(request, 'accounts/rooms/attendance_page.html', {
        'booking': booking,
        'public_status': public_status,
        'step': 'email',
    })




@login_required
@require_POST
def accept_registration_view(request, pk, action):
    """
    POST /accounts/registration/<pk>/accept/  or  /reject/  or  /remove/
    Accepts, rejects, or removes a MeetingAttendee registration.
    Only the booking owner or a superuser can call this.
    """
    attendee = get_object_or_404(MeetingAttendee, pk=pk)
    booking = attendee.booking

    is_owner = booking.requested_by == request.user
    is_room_approver = RoomApprover.objects.filter(
        room=booking.room, user=request.user, is_active=True
    ).exists()

    if not (is_owner or is_room_approver or request.user.is_superuser):
        return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

    if action == 'accept':
        if hasattr(attendee, 'is_accepted'):
            attendee.is_accepted = True
            attendee.save(update_fields=['is_accepted'])
        return JsonResponse({'success': True, 'action': 'accepted', 'pk': pk})

    elif action == 'reject':
        # Delete the registration entirely — setting is_accepted=False would
        # just make it reappear on every refresh as "pending".
        attendee.delete()
        return JsonResponse({'success': True, 'action': 'rejected', 'pk': pk})

    elif action == 'remove':
        attendee.delete()
        return JsonResponse({'success': True, 'action': 'removed', 'pk': pk})

    return JsonResponse({'success': False, 'message': 'Invalid action.'}, status=400)

def meeting_registration_success_view(request):
    return render(request, 'accounts/rooms/meeting_registration_success.html')


# --- FULLY IMPLEMENTED QR CODE VIEW ---
def meeting_qr_code_view(request, registration_code):
    """
    Generates a polished, professional meeting pass card image (PNG) containing:
      - Branded UNPASS header with deep-navy background and teal accents
      - Meeting title (wrapping), room, date, time and live status pill
      - Agency logo (if configured) in the header, or code badge fallback
      - High-error-correction QR code in a framed panel with corner accents
      - Scan instruction and branded footer
    """
    booking = get_object_or_404(
        RoomBooking.objects.select_related('room', 'requested_by', 'requested_by__agency'),
        registration_code=registration_code
    )

    block_reason = _booking_public_link_block_reason(booking)
    if block_reason:
        raise Http404("This public meeting link is no longer available.")

    # Encode the attendance URL in the QR when attendance tracking is on;
    # otherwise encode the standard registration URL.
    if booking.enable_attendance:
        registration_url = request.build_absolute_uri(
            reverse('accounts:meeting_attendance_page', args=[registration_code])
        )
    else:
        registration_url = request.build_absolute_uri(
            reverse('accounts:meeting_registration', args=[registration_code])
        )

    # ── Colour palette ─────────────────────────────────────────────────────
    DEEP_NAVY = (10,  25,  61)
    TEAL      = (0,  150, 199)
    GOLD      = (255, 184,  28)
    WHITE     = (255, 255, 255)
    MID_GREY  = (148, 163, 184)
    DARK_TEXT = (15,  23,  42)
    SOFT_TEXT = (71,  85, 105)

    STATUS_FILL = {
        "live":     (16,  185, 129),
        "open":     TEAL,
        "full":     (245, 158,  11),
        "ended":    (100, 116, 139),
        "disabled": (100, 116, 139),
    }

    # ── Font loader (Poppins → DejaVu → default) ───────────────────────────
    _FONT_CANDIDATES = [
        "/usr/share/fonts/truetype/google-fonts/Poppins-{w}.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans{w}.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans{w}.ttf",
    ]
    _WEIGHT = {
        "Bold":    ("Bold",    "-Bold",    "-Bold"),
        "Medium":  ("Medium",  "-Bold",    "-Bold"),
        "Regular": ("Regular", "",         ""),
        "Light":   ("Light",   "",         ""),
    }

    def load_font(weight, size):
        variants = _WEIGHT.get(weight, ("Regular", "", ""))
        for tpl, v in zip(_FONT_CANDIDATES, variants):
            try:
                return ImageFont.truetype(tpl.format(w=v), size)
            except Exception:
                continue
        return ImageFont.load_default()

    # ── Text wrap helper ───────────────────────────────────────────────────
    def wrap_text(d, text, font, max_w):
        words, lines, cur = text.split(), [], ""
        for word in words:
            trial = (cur + " " + word).strip()
            if d.textbbox((0, 0), trial, font=font)[2] <= max_w:
                cur = trial
            else:
                if cur:
                    lines.append(cur)
                cur = word
        if cur:
            lines.append(cur)
        return lines or [""]

    # ── Duration helper ────────────────────────────────────────────────────
    def fmt_dur(s, e):
        mins = (e.hour * 60 + e.minute) - (s.hour * 60 + s.minute)
        if mins <= 0:
            return ""
        h, m = divmod(mins, 60)
        if h and m:
            return f"  ({h}h {m}m)"
        return f"  ({h}h)" if h else f"  ({m}m)"

    # ── Rounded-rect mask helper ───────────────────────────────────────────
    def rr_mask(size, radius):
        mask = Image.new("L", size, 0)
        ImageDraw.Draw(mask).rounded_rectangle(
            [0, 0, size[0] - 1, size[1] - 1], radius=radius, fill=255
        )
        return mask

    # ── Canvas ─────────────────────────────────────────────────────────────
    W, H     = 900, 1320
    RADIUS   = 40
    MARGIN   = 54
    HEADER_H = 270

    card = Image.new("RGBA", (W, H), (0, 0, 0, 0))

    # White rounded background
    bg = Image.new("RGBA", (W, H), (*WHITE, 255))
    card.paste(bg, mask=rr_mask((W, H), RADIUS))

    # Deep-navy header block (rounded top, flat bottom)
    hdr = Image.new("RGBA", (W, HEADER_H + RADIUS), (*DEEP_NAVY, 255))
    card.paste(hdr, (0, 0), rr_mask((W, HEADER_H + RADIUS), RADIUS))
    card.paste(Image.new("RGBA", (W, RADIUS), (*DEEP_NAVY, 255)), (0, HEADER_H))

    # Decorative translucent teal depth circles
    for cx, cy, r, a in [(810, -35, 210, 16), (860, 148, 118, 10), (28, 238, 72, 14)]:
        circ = Image.new("RGBA", (W, H), (0, 0, 0, 0))
        ImageDraw.Draw(circ).ellipse([cx - r, cy - r, cx + r, cy + r], fill=(*TEAL, a))
        card = Image.alpha_composite(card, circ)

    draw = ImageDraw.Draw(card)

    # ── Brand wordmark ─────────────────────────────────────────────────────
    draw.text((MARGIN, 48),  "UNPASS",              fill=WHITE,           font=load_font("Bold",    56))
    draw.rounded_rectangle([MARGIN, 116, MARGIN + 162, 123], radius=3,  fill=(*TEAL,  255))
    draw.text((MARGIN, 130), "Official Meeting Pass", fill=(*TEAL, 255), font=load_font("Medium",  26))

    # Gold + ghost dots
    draw.ellipse([MARGIN, 180, MARGIN + 22, 202],          fill=GOLD)
    draw.ellipse([MARGIN + 30, 182, MARGIN + 48, 200],     fill=(*WHITE, 85))

    # ── Agency logo or code badge (top-right of header) ───────────────────
    LX1, LY1 = W - 175, 44
    LX2, LY2 = W - 44,  184
    LW, LH   = LX2 - LX1, LY2 - LY1

    agency     = getattr(booking.requested_by, 'agency', None)
    logo_placed = False
    if agency and getattr(agency, 'logo', None):
        try:
            logo_img = Image.open(agency.logo.path).convert("RGBA")
            logo_img.thumbnail((LW - 16, LH - 16), Image.LANCZOS)
            draw.rounded_rectangle([LX1, LY1, LX2, LY2], radius=18, fill=(*WHITE, 230))
            lx = LX1 + (LW - logo_img.width)  // 2
            ly = LY1 + (LH - logo_img.height) // 2
            card.paste(logo_img, (lx, ly), logo_img)
            draw = ImageDraw.Draw(card)
            logo_placed = True
        except Exception:
            pass

    if not logo_placed:
        code_txt = (agency.code if agency else "UN")[:4]
        draw.rounded_rectangle([LX1, LY1, LX2, LY2], radius=18,
                               fill=(*WHITE, 32), outline=(*WHITE, 60), width=2)
        bb = draw.textbbox((0, 0), code_txt, font=load_font("Bold", 36))
        draw.text(
            (LX1 + (LW - (bb[2] - bb[0])) // 2, LY1 + (LH - (bb[3] - bb[1])) // 2),
            code_txt, fill=(*WHITE, 200), font=load_font("Bold", 36)
        )

    # ── Meeting title (wraps to 2 lines) ───────────────────────────────────
    t_font  = load_font("Bold", 30)
    t_lines = wrap_text(draw, booking.title, t_font, LX1 - MARGIN - 16)[:2]
    ty = 205
    for line in t_lines:
        draw.text((MARGIN, ty), line, fill=WHITE, font=t_font)
        ty += 36

    # ── Teal divider stripe ─────────────────────────────────────────────────
    draw.rectangle([0, HEADER_H, W, HEADER_H + 6], fill=TEAL)

    # ── Info rows ───────────────────────────────────────────────────────────
    y      = HEADER_H + 34
    ROW_H  = 80
    ICO_SZ = 34

    def info_row(icon_ch, label_str, value_str, y_pos):
        draw.rounded_rectangle(
            [MARGIN, y_pos + 2, MARGIN + ICO_SZ, y_pos + 2 + ICO_SZ],
            radius=10, fill=(*TEAL, 30)
        )
        draw.text((MARGIN + 5, y_pos + 7), icon_ch, fill=TEAL, font=load_font("Bold", 17))
        draw.text((MARGIN + ICO_SZ + 14, y_pos),      label_str, fill=MID_GREY,  font=load_font("Regular", 17))
        draw.text((MARGIN + ICO_SZ + 14, y_pos + 20), value_str, fill=DARK_TEXT, font=load_font("Bold",    24))

    room_val = f"{booking.room.name}  ·  {booking.room.code}"
    if booking.room.location:
        room_val += f"  ·  {booking.room.location}"

    date_val = booking.date.strftime("%A, %d %B %Y")
    time_val = (
        f"{booking.start_time.strftime('%H:%M')} – "
        f"{booking.end_time.strftime('%H:%M')}"
        + fmt_dur(booking.start_time, booking.end_time)
    )

    info_row("⊞", "ROOM", room_val[:60],  y); y += ROW_H
    info_row("◈", "DATE", date_val,        y); y += ROW_H
    info_row("◷", "TIME", time_val,        y); y += ROW_H

    # Separator rule
    draw.line([MARGIN, y + 4, W - MARGIN, y + 4], fill=(*MID_GREY, 55), width=1)
    y += 22

    # ── Status pill ─────────────────────────────────────────────────────────
    pub_status   = _booking_public_link_status(booking)
    s_color      = STATUS_FILL.get(pub_status["code"], TEAL)
    pill_font    = load_font("Bold", 19)
    pill_text    = f"  ●  {pub_status['label']}  "
    pill_w       = draw.textbbox((0, 0), pill_text, font=pill_font)[2] + 24
    draw.rounded_rectangle([MARGIN, y, MARGIN + pill_w, y + 42], radius=21, fill=(*s_color, 255))
    draw.text((MARGIN + 12, y + 8), pill_text, fill=WHITE, font=pill_font)
    y += 60

    # ── QR code ─────────────────────────────────────────────────────────────
    qr_obj = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=2,
    )
    qr_obj.add_data(registration_url)
    qr_obj.make(fit=True)
    qr_img = qr_obj.make_image(fill_color=DEEP_NAVY, back_color=WHITE).convert("RGBA")
    QR_SZ  = 470
    qr_img = qr_img.resize((QR_SZ, QR_SZ), Image.LANCZOS)

    PAD     = 26
    frame_w = QR_SZ + PAD * 2
    qr_x    = (W - frame_w) // 2
    qr_y    = y + 10

    # Drop shadow
    shadow = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    ImageDraw.Draw(shadow).rounded_rectangle(
        [qr_x - 8, qr_y - 8, qr_x + frame_w + 8, qr_y + frame_w + 8],
        radius=30, fill=(0, 0, 0, 36)
    )
    card = Image.alpha_composite(card, shadow)
    draw = ImageDraw.Draw(card)

    # White frame + teal border
    draw.rounded_rectangle(
        [qr_x, qr_y, qr_x + frame_w, qr_y + frame_w],
        radius=22, fill=WHITE, outline=TEAL, width=4
    )

    # Teal corner accent squares
    csz = 14
    for (cx, cy) in [
        (qr_x + 10,                qr_y + 10),
        (qr_x + frame_w - 10 - csz, qr_y + 10),
        (qr_x + 10,                qr_y + frame_w - 10 - csz),
        (qr_x + frame_w - 10 - csz, qr_y + frame_w - 10 - csz),
    ]:
        draw.rounded_rectangle([cx, cy, cx + csz, cy + csz], radius=4, fill=TEAL)

    card.paste(qr_img, (qr_x + PAD, qr_y + PAD))
    draw = ImageDraw.Draw(card)
    y = qr_y + frame_w + 24

    # ── Scan instruction ────────────────────────────────────────────────────
    draw.text((W // 2, y), "Scan to register or check in",
              fill=SOFT_TEXT, font=load_font("Medium", 22), anchor="mm")
    y += 34
    draw.text((W // 2, y), "Do not share if attendance is restricted",
              fill=MID_GREY, font=load_font("Regular", 17), anchor="mm")

    # ── Footer bar ──────────────────────────────────────────────────────────
    FOOTER_H = 62
    footer_y = H - FOOTER_H

    # Teal rule
    draw.rectangle([0, footer_y - 5, W, footer_y], fill=TEAL)

    # Dark footer (rounded bottom only)
    ft = Image.new("RGBA", (W, FOOTER_H + RADIUS), (*DEEP_NAVY, 255))
    card.paste(ft, (0, footer_y - RADIUS), rr_mask((W, FOOTER_H + RADIUS), RADIUS))
    card.paste(Image.new("RGBA", (W, RADIUS), (*DEEP_NAVY, 255)), (0, footer_y))
    draw = ImageDraw.Draw(card)
    draw.text(
        (W // 2, footer_y + FOOTER_H // 2 - 4),
        "UNDP  ·  Powered by UNPASS",
        fill=(*TEAL, 210), font=load_font("Regular", 18), anchor="mm"
    )

    # ── Composite to RGB and emit ───────────────────────────────────────────
    out = Image.new("RGB", (W, H), WHITE)
    out.paste(card, mask=card.split()[3])

    buffer = io.BytesIO()
    out.save(buffer, format="PNG", dpi=(144, 144))
    buffer.seek(0)

    safe_name = slugify(booking.title)[:40] or "meeting"
    response  = HttpResponse(buffer.getvalue(), content_type="image/png")
    response["Content-Disposition"] = (
        f'inline; filename="unpass_qr_{safe_name}_{booking.pk}.png"'
    )
    return response

@login_required
def meeting_qr_code_download_view(request, registration_code):
    response = meeting_qr_code_view(request, registration_code)
    response["Content-Disposition"] = f'attachment; filename="unpass_meeting_qr_{registration_code}.png"'
    return response

@login_required
@require_GET
def agenda_document_qr_view(request, pk):
    """
    Generates a QR code card image (PNG) that encodes the direct download URL
    of the booking's agenda document.  Attendees scan this to download the file.

    The card uses the same UNPASS design language as meeting_qr_code_view but
    is clearly labelled as an AGENDA / FILE download QR.
    """
    booking = get_object_or_404(
        RoomBooking.objects.select_related('room', 'requested_by', 'requested_by__agency'),
        pk=pk,
    )

    if not booking.agenda_document:
        raise Http404("No agenda document attached to this booking.")

    # Build the full absolute download URL for the agenda file
    file_url = request.build_absolute_uri(booking.agenda_document.url)
    file_name = booking.agenda_document.name.split('/')[-1]

    # ── Colour palette (same as UNPASS brand) ────────────────────────────────
    DEEP_NAVY = (10,  25,  61)
    TEAL      = (0,  150, 199)
    GOLD      = (255, 184,  28)
    WHITE     = (255, 255, 255)
    MID_GREY  = (148, 163, 184)
    DARK_TEXT = (15,  23,  42)
    SOFT_TEXT = (71,  85, 105)
    EMERALD   = (16,  185, 129)   # accent for "document" flavour

    # ── Font loader ───────────────────────────────────────────────────────────
    _FC = [
        "/usr/share/fonts/truetype/google-fonts/Poppins-{w}.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans{w}.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans{w}.ttf",
    ]
    _WM = {
        "Bold":    ("Bold",    "-Bold",    "-Bold"),
        "Medium":  ("Medium",  "-Bold",    "-Bold"),
        "Regular": ("Regular", "",         ""),
    }

    def lf(weight, size):
        for tpl, v in zip(_FC, _WM.get(weight, ("Regular", "", ""))):
            try:
                return ImageFont.truetype(tpl.format(w=v), size)
            except Exception:
                continue
        return ImageFont.load_default()

    def wrap(d, text, font, max_w):
        words, lines, cur = text.split(), [], ""
        for w in words:
            t = (cur + " " + w).strip()
            if d.textbbox((0, 0), t, font=font)[2] <= max_w:
                cur = t
            else:
                if cur: lines.append(cur)
                cur = w
        if cur: lines.append(cur)
        return lines or [""]

    def rr_mask(size, r):
        m = Image.new("L", size, 0)
        ImageDraw.Draw(m).rounded_rectangle([0, 0, size[0]-1, size[1]-1], radius=r, fill=255)
        return m

    # ── Canvas ────────────────────────────────────────────────────────────────
    W, H     = 900, 1160
    RADIUS   = 40
    MG       = 54
    HEADER_H = 248

    card = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    card.paste(Image.new("RGBA", (W, H), (*WHITE, 255)), mask=rr_mask((W, H), RADIUS))

    # Header — deep navy
    hdr = Image.new("RGBA", (W, HEADER_H + RADIUS), (*DEEP_NAVY, 255))
    card.paste(hdr, (0, 0), rr_mask((W, HEADER_H + RADIUS), RADIUS))
    card.paste(Image.new("RGBA", (W, RADIUS), (*DEEP_NAVY, 255)), (0, HEADER_H))

    # Decorative emerald circles (doc-flavour tint)
    for cx, cy, r, a in [(810, -30, 200, 14), (860, 140, 110, 9), (30, 228, 68, 12)]:
        c2 = Image.new("RGBA", (W, H), (0, 0, 0, 0))
        ImageDraw.Draw(c2).ellipse([cx-r, cy-r, cx+r, cy+r], fill=(*EMERALD, a))
        card = Image.alpha_composite(card, c2)

    draw = ImageDraw.Draw(card)

    # Brand wordmark
    draw.text((MG, 44), "UNPASS", fill=WHITE, font=lf("Bold", 52))
    draw.rounded_rectangle([MG, 108, MG+148, 115], radius=3, fill=(*EMERALD, 255))
    draw.text((MG, 122), "Agenda & File Download", fill=(*EMERALD, 255), font=lf("Medium", 24))

    # Gold + ghost dots
    draw.ellipse([MG, 170, MG+20, 190], fill=GOLD)
    draw.ellipse([MG+28, 172, MG+44, 188], fill=(*WHITE, 80))

    # Agency badge top-right
    LX1, LY1 = W-165, 40; LX2, LY2 = W-44, 168; LW, LH = LX2-LX1, LY2-LY1
    agency = getattr(booking.requested_by, 'agency', None)
    logo_placed = False
    if agency and getattr(agency, 'logo', None):
        try:
            logo_img = Image.open(agency.logo.path).convert("RGBA")
            logo_img.thumbnail((LW-14, LH-14), Image.LANCZOS)
            draw.rounded_rectangle([LX1, LY1, LX2, LY2], radius=16, fill=(*WHITE, 220))
            lx = LX1 + (LW - logo_img.width) // 2
            ly = LY1 + (LH - logo_img.height) // 2
            card.paste(logo_img, (lx, ly), logo_img)
            draw = ImageDraw.Draw(card)
            logo_placed = True
        except Exception:
            pass
    if not logo_placed:
        code_txt = (agency.code if agency else "UN")[:4]
        draw.rounded_rectangle([LX1, LY1, LX2, LY2], radius=16,
                               fill=(*WHITE, 28), outline=(*WHITE, 55), width=2)
        bb = draw.textbbox((0, 0), code_txt, font=lf("Bold", 32))
        draw.text(
            (LX1 + (LW-(bb[2]-bb[0]))//2, LY1 + (LH-(bb[3]-bb[1]))//2),
            code_txt, fill=(*WHITE, 190), font=lf("Bold", 32)
        )

    # Meeting title (wraps to 2 lines)
    t_font  = lf("Bold", 28)
    t_lines = wrap(draw, booking.title, t_font, LX1 - MG - 14)[:2]
    ty = 196
    for line in t_lines:
        draw.text((MG, ty), line, fill=WHITE, font=t_font)
        ty += 34

    # Emerald stripe divider
    draw.rectangle([0, HEADER_H, W, HEADER_H + 6], fill=EMERALD)

    # ── Info rows ─────────────────────────────────────────────────────────────
    y = HEADER_H + 28; RSZ = 72; ICO = 32

    def ir(ic, lab, val, yp):
        draw.rounded_rectangle([MG, yp+2, MG+ICO, yp+2+ICO], radius=9, fill=(*EMERALD, 28))
        draw.text((MG+5, yp+7), ic, fill=EMERALD, font=lf("Bold", 16))
        draw.text((MG+ICO+12, yp),    lab, fill=MID_GREY,  font=lf("Regular", 16))
        draw.text((MG+ICO+12, yp+18), val, fill=DARK_TEXT, font=lf("Bold",    22))

    room_val = f"{booking.room.name}  ·  {booking.room.code}"
    date_val = booking.date.strftime("%A, %d %B %Y")
    mins = (booking.end_time.hour*60+booking.end_time.minute) - (booking.start_time.hour*60+booking.start_time.minute)
    h, m = divmod(max(mins, 0), 60)
    dur = f"  ({h}h {m}m)" if h and m else (f"  ({h}h)" if h else f"  ({m}m)" if m else "")
    time_val = f"{booking.start_time.strftime('%H:%M')} – {booking.end_time.strftime('%H:%M')}{dur}"

    ir("⊞", "ROOM", room_val[:58],  y); y += RSZ
    ir("◈", "DATE", date_val,        y); y += RSZ

    # Separator
    draw.line([MG, y+4, W-MG, y+4], fill=(*MID_GREY, 50), width=1); y += 20

    # File name pill
    fn_font = lf("Bold", 18)
    fn_text = f"  📄  {file_name[:55]}  "
    fn_w = draw.textbbox((0, 0), fn_text, font=fn_font)[2] + 22
    draw.rounded_rectangle([MG, y, MG+fn_w, y+40], radius=20, fill=(*EMERALD, 255))
    draw.text((MG+11, y+8), fn_text, fill=WHITE, font=fn_font); y += 56

    # ── QR code ───────────────────────────────────────────────────────────────
    qr_obj = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_H,
                           box_size=10, border=2)
    qr_obj.add_data(file_url)
    qr_obj.make(fit=True)
    qr_img = qr_obj.make_image(fill_color=DEEP_NAVY, back_color=WHITE).convert("RGBA")
    QS = 440
    qr_img = qr_img.resize((QS, QS), Image.LANCZOS)

    PAD = 24; fw = QS + PAD*2; qx = (W-fw)//2; qy = y+8

    shadow = Image.new("RGBA", (W, H), (0,0,0,0))
    ImageDraw.Draw(shadow).rounded_rectangle([qx-6, qy-6, qx+fw+6, qy+fw+6],
                                             radius=28, fill=(0,0,0,32))
    card = Image.alpha_composite(card, shadow)
    draw = ImageDraw.Draw(card)

    draw.rounded_rectangle([qx, qy, qx+fw, qy+fw], radius=20, fill=WHITE,
                           outline=EMERALD, width=4)
    csz = 13
    for cx2, cy2 in [(qx+9,qy+9),(qx+fw-9-csz,qy+9),(qx+9,qy+fw-9-csz),(qx+fw-9-csz,qy+fw-9-csz)]:
        draw.rounded_rectangle([cx2,cy2,cx2+csz,cy2+csz], radius=4, fill=EMERALD)

    card.paste(qr_img, (qx+PAD, qy+PAD))
    draw = ImageDraw.Draw(card)
    y = qy + fw + 22

    # Instructions
    draw.text((W//2, y), "Scan to download the agenda / document",
              fill=SOFT_TEXT, font=lf("Medium", 21), anchor="mm"); y += 32
    draw.text((W//2, y), "File download will open in your browser",
              fill=MID_GREY, font=lf("Regular", 16), anchor="mm")

    # ── Footer ────────────────────────────────────────────────────────────────
    FH = 58; fy = H - FH
    draw.rectangle([0, fy-4, W, fy], fill=EMERALD)
    ft = Image.new("RGBA", (W, FH+RADIUS), (*DEEP_NAVY, 255))
    card.paste(ft, (0, fy-RADIUS), rr_mask((W, FH+RADIUS), RADIUS))
    card.paste(Image.new("RGBA", (W, RADIUS), (*DEEP_NAVY, 255)), (0, fy))
    draw = ImageDraw.Draw(card)
    draw.text((W//2, fy+FH//2-4), "EasyOffice  ·  Powered by UNPASS",
              fill=(*EMERALD, 200), font=lf("Regular", 17), anchor="mm")

    out = Image.new("RGB", (W, H), WHITE)
    out.paste(card, mask=card.split()[3])

    buf = io.BytesIO()
    out.save(buf, format="PNG", dpi=(144, 144))
    buf.seek(0)

    safe = slugify(booking.title)[:35] or "meeting"
    resp = HttpResponse(buf.getvalue(), content_type="image/png")
    resp["Content-Disposition"] = f'inline; filename="agenda_qr_{safe}_{pk}.png"'
    return resp



@login_required
def booking_attendance_export_csv(request, pk):
    booking = get_object_or_404(
        RoomBooking.objects.prefetch_related('attendance_records', 'attendees'),
        pk=pk
    )

    filename = f"{slugify(booking.title)}_attendance.csv"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        'Meeting Title',
        'Date',
        'Room',
        'Mode',
        'Name',
        'Email',
        'Organization',
        'Status',
        'Invited',
        'Pre-registered',
        'Checked In At',
    ])

    if booking.enable_attendance:
        rows = booking.attendance_records.all().order_by('checked_in_at')
        for r in rows:
            writer.writerow([
                booking.title,
                booking.date,
                booking.room.name,
                'Attendance',
                r.name,
                r.email,
                r.organization,
                r.status,
                'Yes' if r.was_invited else 'No',
                'Yes' if r.was_preregistered else 'No',
                timezone.localtime(r.checked_in_at).strftime('%Y-%m-%d %H:%M:%S') if r.checked_in_at else '',
            ])
    else:
        rows = booking.attendees.all().order_by('registered_at')
        for a in rows:
            writer.writerow([
                booking.title,
                booking.date,
                booking.room.name,
                'Registration',
                a.name,
                a.email,
                a.organization,
                'registered',
                '',
                '',
                timezone.localtime(a.registered_at).strftime('%Y-%m-%d %H:%M:%S') if a.registered_at else '',
            ])

    return response


@login_required
def booking_attendance_export_excel(request, pk):
    booking = get_object_or_404(
        RoomBooking.objects.select_related('room', 'requested_by', 'requested_by__agency')
        .prefetch_related('attendance_records', 'attendees'),
        pk=pk
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Column widths
    widths = {
        "A": 22, "B": 16, "C": 22, "D": 18, "E": 24,
        "F": 30, "G": 24, "H": 18, "I": 12, "J": 15, "K": 22
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Styles
    header_fill = PatternFill("solid", fgColor="009EDB")
    title_fill = PatternFill("solid", fgColor="133B5C")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9E5EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header panel
    ws.merge_cells("A1:K1")
    ws["A1"] = "UNPASS Meeting Attendance Export"
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A1"].alignment = center

    ws.merge_cells("A2:K2")
    ws["A2"] = booking.title
    ws["A2"].font = Font(bold=True, size=13)
    ws["A2"].alignment = center

    ws["A4"] = "Date"
    ws["B4"] = str(booking.date)
    ws["D4"] = "Room"
    ws["E4"] = booking.room.name
    ws["G4"] = "Mode"
    ws["H4"] = "Attendance" if booking.enable_attendance else "Registration"
    ws["J4"] = "Status"
    ws["K4"] = _booking_public_link_status(booking)["label"]

    for cell in ["A4", "D4", "G4", "J4"]:
        ws[cell].font = bold_font

    headers = [
        "Name", "Email", "Organization", "Status", "Invited",
        "Pre-registered", "Timestamp", "Meeting", "Room", "Date", "Source"
    ]
    row_num = 6
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=idx, value=header)
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = center
        cell.border = border

    row_num += 1

    if booking.enable_attendance:
        rows = booking.attendance_records.all().order_by('checked_in_at')
        for r in rows:
            values = [
                r.name,
                r.email,
                r.organization,
                r.status,
                "Yes" if r.was_invited else "No",
                "Yes" if r.was_preregistered else "No",
                timezone.localtime(r.checked_in_at).strftime("%Y-%m-%d %H:%M:%S") if r.checked_in_at else "",
                booking.title,
                booking.room.name,
                str(booking.date),
                "Attendance Record",
            ]
            for idx, val in enumerate(values, start=1):
                c = ws.cell(row=row_num, column=idx, value=val)
                c.border = border
            row_num += 1
    else:
        rows = booking.attendees.all().order_by('registered_at')
        for a in rows:
            values = [
                a.name,
                a.email,
                a.organization,
                "registered",
                "",
                "",
                timezone.localtime(a.registered_at).strftime("%Y-%m-%d %H:%M:%S") if a.registered_at else "",
                booking.title,
                booking.room.name,
                str(booking.date),
                "Meeting Registration",
            ]
            for idx, val in enumerate(values, start=1):
                c = ws.cell(row=row_num, column=idx, value=val)
                c.border = border
            row_num += 1

    filename = f"{slugify(booking.title)}_attendance.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
@require_POST
def cancel_booking(request, pk):
    """
    Cancel a single booking.
    Only the requester can cancel their own bookings.
    Only pending or approved bookings can be cancelled.
    """
    booking = get_object_or_404(RoomBooking, pk=pk)

    # Check permission - only requester can cancel
    if booking.requested_by != request.user:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'You can only cancel your own bookings.'
            }, status=403)
        else:
            messages.error(request, 'You can only cancel your own bookings.')
            return redirect('accounts:my_bookings')

    # Check if booking can be cancelled
    if booking.status not in ['pending', 'approved']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': f'Cannot cancel a {booking.status} booking.'
            }, status=400)
        else:
            messages.error(request, f'Cannot cancel a {booking.status} booking.')
            return redirect('accounts:my_bookings')

    # Check if part of a series
    if booking.series:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'This booking is part of a recurring series. Please cancel the entire series instead.'
            }, status=400)
        else:
            messages.error(request,
                           'This booking is part of a recurring series. Please cancel the entire series instead.')
            return redirect('accounts:my_bookings')

    # Store info for notification
    room_name = booking.room.name
    booking_title = booking.title
    booking_date = booking.date
    was_approved = booking.status == 'approved'

    # Cancel the booking
    booking.status = 'cancelled'
    booking.save(update_fields=['status'])

    # Send notification email to approvers if it was approved
    if was_approved:
        notify_approvers_booking_cancelled(booking)

    # Return response
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': f'Booking cancelled successfully.',
            'booking_id': booking.id
        })
    else:
        messages.success(
            request,
            f'Booking "{booking_title}" for {room_name} on {booking_date} has been cancelled.'
        )
        return redirect('accounts:my_bookings')


@login_required
@require_POST
def cancel_booking_series(request, pk):
    """
    Cancel an entire recurring booking series.
    Only the requester can cancel their own series.
    Only pending or approved series can be cancelled.
    """
    series = get_object_or_404(
        RoomBookingSeries.objects.select_related('room')
        .annotate(occurrence_count=Count('occurrences')),
        pk=pk
    )

    # Check permission - only requester can cancel
    if series.requested_by != request.user:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'You can only cancel your own bookings.'
            }, status=403)
        else:
            messages.error(request, 'You can only cancel your own bookings.')
            return redirect('accounts:my_bookings')

    # Check if series can be cancelled
    if series.status not in ['pending', 'approved']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': f'Cannot cancel a {series.status} series.'
            }, status=400)
        else:
            messages.error(request, f'Cannot cancel a {series.status} series.')
            return redirect('accounts:my_bookings')

    # Store info for notification
    room_name = series.room.name
    series_title = series.title
    occurrence_count = series.occurrence_count
    was_approved = series.status == 'approved'

    # Cancel the series and all its occurrences
    with transaction.atomic():
        series.status = 'cancelled'
        series.save(update_fields=['status'])

        # Cancel all occurrences
        series.occurrences.update(status='cancelled')

    # Send notification email to approvers if it was approved
    if was_approved:
        notify_approvers_series_cancelled(series, occurrence_count)

    # Return response
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': f'Series with {occurrence_count} occurrences cancelled successfully.',
            'series_id': series.id,
            'occurrence_count': occurrence_count
        })
    else:
        messages.success(
            request,
            f'Recurring series "{series_title}" for {room_name} with {occurrence_count} occurrences has been cancelled.'
        )
        return redirect('accounts:my_bookings')


@login_required
@require_POST
def cancel_series_occurrence(request, pk):
    """
    Cancel a single occurrence within a recurring series.
    Only the requester can cancel their own bookings.
    Only pending or approved occurrences can be cancelled.
    """
    occurrence = get_object_or_404(
        RoomBooking.objects.select_related('room', 'series', 'series__requested_by'),
        pk=pk
    )

    # Check if this booking is part of a series
    if not occurrence.series:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'This is not part of a recurring series. Use the regular cancel function.'
            }, status=400)
        else:
            messages.error(request, 'This is not part of a recurring series.')
            return redirect('accounts:my_bookings')

    # Check permission - only requester can cancel
    if occurrence.series.requested_by != request.user:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'You can only cancel your own bookings.'
            }, status=403)
        else:
            messages.error(request, 'You can only cancel your own bookings.')
            return redirect('accounts:my_bookings')

    # Check if occurrence can be cancelled
    if occurrence.status not in ['pending', 'approved']:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': f'Cannot cancel a {occurrence.status} occurrence.'
            }, status=400)
        else:
            messages.error(request, f'Cannot cancel a {occurrence.status} occurrence.')
            return redirect('accounts:my_bookings')

    # Store info for notification
    room_name = occurrence.room.name
    occurrence_date = occurrence.date
    occurrence_time = f"{occurrence.start_time} – {occurrence.end_time}"
    series_title = occurrence.series.title
    was_approved = occurrence.status == 'approved'

    # Cancel the occurrence
    occurrence.status = 'cancelled'
    occurrence.save(update_fields=['status'])

    # Check if all occurrences are now cancelled
    series = occurrence.series
    active_occurrences = series.occurrences.exclude(status='cancelled').count()

    # If no active occurrences left, mark series as cancelled too
    if active_occurrences == 0:
        series.status = 'cancelled'
        series.save(update_fields=['status'])

    # Send notification email to approvers if it was approved
    if was_approved:
        notify_approvers_occurrence_cancelled(occurrence)

    # Return response
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'message': f'Occurrence on {occurrence_date} cancelled successfully.',
            'occurrence_id': occurrence.id,
            'occurrence_date': str(occurrence_date),
            'active_occurrences': active_occurrences,
            'series_cancelled': active_occurrences == 0
        })
    else:
        messages.success(
            request,
            f'Occurrence of "{series_title}" on {occurrence_date} has been cancelled. '
            f'The room is now available for this time slot.'
        )
        return redirect('accounts:my_bookings')


@method_decorator(staff_member_required, name='dispatch')
class RoomCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """Create a new room (superuser only)."""
    model = Room
    form_class = RoomForm
    template_name = "accounts/rooms/room_form.html"
    success_url = reverse_lazy("accounts:room_list")

    def test_func(self):
        return self.request.user.is_superuser

    def form_valid(self, form):
        messages.success(self.request, f"Room '{form.instance.name}' created successfully!")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = 'Add New Room'
        ctx['submit_text'] = 'Create Room'
        return ctx


@method_decorator(staff_member_required, name='dispatch')
class RoomUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Edit an existing room (superuser only)."""
    model = Room
    form_class = RoomForm
    template_name = "accounts/rooms/room_form.html"

    def test_func(self):
        return self.request.user.is_superuser

    def form_valid(self, form):
        messages.success(self.request, f"Room '{form.instance.name}' updated successfully!")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse("accounts:room_list")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['page_title'] = f'Edit Room: {self.object.name}'
        ctx['submit_text'] = 'Update Room'
        ctx['is_edit'] = True
        return ctx


@staff_member_required
def room_delete_view(request, pk):
    """Delete/deactivate a room (superuser only)."""
    if not request.user.is_superuser:
        messages.error(request, "You don't have permission to delete rooms.")
        return redirect("accounts:room_list")

    room = get_object_or_404(Room, pk=pk)

    if request.method == "POST":
        room_name = room.name
        # Soft delete - just deactivate
        room.is_active = False
        room.save()
        messages.warning(request, f"Room '{room_name}' has been deactivated.")
        return redirect("accounts:room_list")

    return render(request, "accounts/rooms/room_confirm_delete.html", {"room": room})


def _parse_iso_dt(value: str):
    """
    FullCalendar sends ISO datetimes like:
      2026-02-19T00:00:00Z
      2026-02-19T00:00:00+00:00
      2026-02-19T00:00:00
    """
    if not value:
        return None

    value = value.strip()

    # Handle Zulu time (Z)
    if value.endswith("Z"):
        value = value[:-1] + "+00:00"

    try:
        dt = datetime.fromisoformat(value)
    except ValueError:
        return None

    # If naive, assume current timezone
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())

    return dt


@login_required
def room_bookings_calendar(request):
    """
    Calendar page (Day/Week/Month/List).
    Loads bookings via AJAX from room_bookings_events endpoint.
    """
    rooms = Room.objects.filter(is_active=True).order_by("name")

    # Optional: restrict rooms per agency if you have agency field
    # if hasattr(Room, "agency_id") and request.user.agency_id:
    #     rooms = rooms.filter(agency_id=request.user.agency_id)

    context = {
        "rooms": rooms,
        "status_choices": RoomBooking.STATUS_CHOICES,
    }
    return render(request, "accounts/rooms/room_bookings_calendar.html", context)


@require_GET
@login_required
def room_bookings_events(request):
    """
    JSON endpoint for FullCalendar.
    Query params:
      start=ISO datetime
      end=ISO datetime
      mine=1 (optional)
      room_id= (optional)
      status= (optional)
      future_only=1 (optional default on frontend)
    """
    start_dt = _parse_iso_dt(request.GET.get("start"))
    end_dt = _parse_iso_dt(request.GET.get("end"))

    mine = (request.GET.get("mine") or "").strip() == "1"
    room_id = (request.GET.get("room_id") or "").strip()
    status = (request.GET.get("status") or "").strip()
    future_only = (request.GET.get("future_only") or "1").strip() == "1"

    # Base queryset
    qs = RoomBooking.objects.select_related("room", "requested_by").all()

    # Optional: agency restriction if your models have agency
    # if hasattr(RoomBooking, "agency_id") and request.user.agency_id:
    #     qs = qs.filter(room__agency_id=request.user.agency_id)

    if mine:
        qs = qs.filter(requested_by=request.user)

    if room_id.isdigit():
        qs = qs.filter(room_id=int(room_id))

    if status:
        qs = qs.filter(status=status)

    # Date range from FullCalendar
    if start_dt and end_dt:
        qs = qs.filter(date__gte=start_dt.date(), date__lte=end_dt.date())

    # Today/future only (default ON)
    if future_only:
        qs = qs.filter(date__gte=timezone.localdate())

    # Build events list
    events = []
    tz = timezone.get_current_timezone()

    for b in qs:
        start = datetime.combine(b.date, b.start_time)
        end = datetime.combine(b.date, b.end_time)

        if timezone.is_naive(start):
            start = timezone.make_aware(start, tz)
        if timezone.is_naive(end):
            end = timezone.make_aware(end, tz)

        # Status color classes (FullCalendar can use classNames)
        # You can style these in CSS.
        class_names = [f"bk-status-{b.status}"]

        events.append({
            "id": b.id,
            "title": f"{b.room.name} • {b.title}",
            "start": start.isoformat(),
            "end": end.isoformat(),
            "allDay": False,
            "classNames": class_names,
            "extendedProps": {
                "room": b.room.name,
                "status": b.status,
                "requested_by": (b.requested_by.get_full_name() or b.requested_by.username),
                "description": b.description or "",
            }
        })

    return JsonResponse(events, safe=False)


@login_required
@require_GET
def check_availability_api(request):
    """
    Returns JSON:
      { available: true }
      { available: false, conflict: { start, end, title } }
      { available: false, conflict: {...}, next_slot: { start, end, date, label } }  # when find_next=1
    """
    room_id = request.GET.get('room', '').strip()
    date_str = request.GET.get('date', '').strip()
    start_str = request.GET.get('start', '').strip()
    end_str = request.GET.get('end', '').strip()
    find_next = request.GET.get('find_next', '0') == '1'
    duration = int(request.GET.get('duration', 60) or 60)

    # Validate inputs
    if not all([room_id, date_str, start_str, end_str]):
        return JsonResponse({'available': True})  # incomplete — don't block

    try:
        room = Room.objects.get(pk=room_id, is_active=True)
        check_date = date_cls.fromisoformat(date_str)
        start_time = time_cls.fromisoformat(start_str)
        end_time = time_cls.fromisoformat(end_str)
    except (Room.DoesNotExist, ValueError):
        return JsonResponse({'available': True})

    if end_time <= start_time:
        return JsonResponse({'available': True})

    # Look for overlapping bookings
    overlap_qs = RoomBooking.objects.filter(
        room=room,
        date=check_date,
        status__in=('approved', 'pending'),
        start_time__lt=end_time,
        end_time__gt=start_time,
    ).order_by('start_time')

    conflict_booking = overlap_qs.first()

    if not conflict_booking:
        return JsonResponse({'available': True})

    # Build conflict info
    conflict_info = {
        'start': conflict_booking.start_time.strftime('%H:%M'),
        'end': conflict_booking.end_time.strftime('%H:%M'),
        'title': conflict_booking.title if not conflict_booking.requested_by == request.user
        else None,  # hide title if it's their own booking (edge case)
    }

    response = {'available': False, 'conflict': conflict_info}

    if find_next:
        # Compute requested duration
        req_duration = int(
            (datetime.combine(check_date, end_time) - datetime.combine(check_date, start_time))
            .total_seconds() / 60
        )
        next_slot = find_next_available_slot(
            room=room,
            on_date=check_date,
            duration_minutes=req_duration,
            after_time=conflict_booking.end_time,  # start searching after the conflict ends
        )
        response['next_slot'] = next_slot  # None or dict

    return JsonResponse(response)

@login_required
@require_POST
def toggle_booking_option_view(request, pk):
    """
    Toggle meeting options directly from booking detail page.
    Only the booking owner or a superuser can change these options.
    """
    booking = get_object_or_404(RoomBooking, pk=pk)

    if request.user != booking.requested_by and not request.user.is_superuser:
        return JsonResponse({
            "success": False,
            "message": "You are not allowed to update this booking."
        }, status=403)

    field = (request.POST.get("field") or "").strip()
    raw_value = (request.POST.get("value") or "").strip().lower()

    allowed_boolean_fields = {"enable_attendance", "enable_invite_link", "auto_accept_registration"}

    if field in allowed_boolean_fields:
        value = raw_value in {"1", "true", "yes", "on"}
        setattr(booking, field, value)
        booking.save(update_fields=[field])

        # If invite link is turned on and no code exists yet, create one
        if field == "enable_invite_link" and value and not booking.registration_code:
            import uuid
            booking.registration_code = uuid.uuid4()
            booking.save(update_fields=["registration_code"])

        public_status = _booking_public_link_status(booking)

        return JsonResponse({
            "success": True,
            "field": field,
            "value": getattr(booking, field),
            "public_status": public_status,
            "message": f"{field.replace('_', ' ').title()} updated successfully."
        })

    return JsonResponse({
        "success": False,
        "message": "Invalid option selected."
    }, status=400)